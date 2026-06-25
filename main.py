#!/usr/bin/env python3
"""
Instagram Account Buyer Bot
Production-ready Telegram bot for Railway + GitHub deployment.
"""

import asyncio
import json
import logging
import os
import random
import re
import string
import tempfile
from datetime import datetime, timezone
import hashlib
import hmac
import base64
import struct
import time
import firebase_admin
import openpyxl
from firebase_admin import credentials, db
from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    Update,
)
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

# ─────────────────────────────────────────────
# Logging
# ─────────────────────────────────────────────
logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# Environment Variables
# ─────────────────────────────────────────────
BOT_TOKEN = os.environ["BOT_TOKEN"]
FIREBASE_CONFIG = os.environ["FIREBASE_CONFIG"]
REQUIRED_CHANNEL = os.environ.get("REQUIRED_CHANNEL", "")
CHANNEL_LINK = os.environ.get("CHANNEL_LINK", "")
FIREBASE_DATABASE_URL = os.environ["FIREBASE_DATABASE_URL"]

# ─────────────────────────────────────────────
# Admin IDs
# ─────────────────────────────────────────────
ADMIN_IDS = {8907284640}

# Tracks pending auto-approve asyncio tasks: "{user_id}_{sub_id}" -> Task
_pending_auto_approve: dict = {}
# Exchange rate for BDT (1 USD = 120 BDT)
USD_TO_BDT_RATE = 120.0
BOT_ENABLED = True
# ─────────────────────────────────────────────
# Conversation States
# ─────────────────────────────────────────────
(
    HOME,
    TASK_MENU,
    TASK_2FA_INFO,
    TASK_2FA_STARTED,
    TASK_2FA_AWAIT_KEY,
    TASK_2FA_1H_INFO,
    TASK_2FA_1H_STARTED,
    TASK_2FA_1H_AWAIT_KEY,
    WITHDRAW_MENU,
    WITHDRAW_AMOUNT,
    WITHDRAW_ADDRESS,
    ADMIN_ACTS_VIEW,
    TASK_FB_INFO,
    TASK_FB_AWAIT_UID,
    TASK_FB_AWAIT_COOKIES,
    TASK_FB_STARTED,
    ADMIN_FBACTS_VIEW,
) = range(17)

# ─────────────────────────────────────────────
# Firebase Initialisation
# ─────────────────────────────────────────────
def init_firebase():
    config = json.loads(FIREBASE_CONFIG)
    config.pop("databaseURL", None)
    database_url = FIREBASE_DATABASE_URL.strip()
    if not database_url:
        raise ValueError("FIREBASE_DATABASE_URL environment variable is missing or empty.")
    cred = credentials.Certificate(config)
    firebase_admin.initialize_app(cred, {"databaseURL": database_url})
    logger.info("Firebase initialised successfully.")

# ─────────────────────────────────────────────
# Referral Helpers
# ─────────────────────────────────────────────
def get_referral_data(user_id: int) -> dict:
    try:
        data = db.reference(f"referrals/{user_id}").get()
        if data is None:
            data = {
                "referral_code": str(user_id),
                "referred_by": None,
                "referrals": [],
                "total_earned": 0.0,
                "joined_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
            }
            db.reference(f"referrals/{user_id}").set(data)
        return data
    except Exception as e:
        logger.error(f"get_referral_data({user_id}) failed: {e}")
        return {"referral_code": str(user_id), "referred_by": None, "referrals": [], "total_earned": 0.0}


def update_referral_data(user_id: int, updates: dict):
    try:
        db.reference(f"referrals/{user_id}").update(updates)
    except Exception as e:
        logger.error(f"update_referral_data({user_id}) failed: {e}")


def add_referral(user_id: int, referrer_id: int):
    if user_id == referrer_id:
        return False
    try:
        referrer_data = get_referral_data(referrer_id)
        user_data = get_referral_data(user_id)
        if user_data.get("referred_by") is not None:
            return False
        update_referral_data(user_id, {"referred_by": referrer_id})
        referrals = referrer_data.get("referrals", [])
        if user_id not in referrals:
            referrals.append(user_id)
            update_referral_data(referrer_id, {"referrals": referrals})
        logger.info(f"Referral added: {user_id} referred by {referrer_id}")
        return True
    except Exception as e:
        logger.error(f"add_referral({user_id}, {referrer_id}) failed: {e}")
        return False


def get_referral_stats(user_id: int) -> dict:
    data = get_referral_data(user_id)
    referrals = data.get("referrals", [])
    now = datetime.now(timezone.utc)
    new_last_24h = 0
    for ref_id in referrals:
        ref_data = get_referral_data(ref_id)
        joined_at_str = ref_data.get("joined_at", "")
        if joined_at_str:
            try:
                joined_at = datetime.strptime(joined_at_str, "%Y-%m-%d %H:%M:%S UTC").replace(tzinfo=timezone.utc)
                if (now - joined_at).total_seconds() <= 24 * 3600:
                    new_last_24h += 1
            except Exception:
                pass
    return {
        "total": len(referrals),
        "new_last_24h": new_last_24h,
        "total_earned": data.get("total_earned", 0.0)
    }

# ─────────────────────────────────────────────
# Bot State
# ─────────────────────────────────────────────
def get_bot_state() -> dict:
    try:
        state = db.reference("bot/state").get()
        if state is None:
            state = {
                "enabled": True,
                "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
                "updated_by": None
            }
            db.reference("bot/state").set(state)
        return state
    except Exception as e:
        logger.error(f"get_bot_state failed: {e}")
        return {"enabled": True, "last_updated": "", "updated_by": None}


def set_bot_state(enabled: bool, admin_id: int = None):
    try:
        state = {
            "enabled": enabled,
            "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            "updated_by": admin_id
        }
        db.reference("bot/state").set(state)
        logger.info(f"Bot state changed to {'ON' if enabled else 'OFF'} by admin {admin_id}")
    except Exception as e:
        logger.error(f"set_bot_state failed: {e}")

# ─────────────────────────────────────────────
# Leaderboard Helpers
# ─────────────────────────────────────────────
def get_today_stats() -> dict:
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    all_subs = get_all_submissions()
    today_stats = {}
    for uid, subs in all_subs.items():
        if subs:
            today_count = 0
            for sub_id, sub_data in subs.items():
                sub_date = sub_data.get('datetime', '')
                if sub_date.startswith(today):
                    today_count += 1
            if today_count > 0:
                today_stats[uid] = today_count
    return today_stats


def get_leaderboard_data() -> dict:
    try:
        data = db.reference("leaderboard/data").get()
        return data or {}
    except Exception as e:
        logger.error(f"get_leaderboard_data failed: {e}")
        return {}


def set_leaderboard_data(data: dict):
    try:
        db.reference("leaderboard/data").set(data)
    except Exception as e:
        logger.error(f"set_leaderboard_data failed: {e}")


def get_leaderboard_settings() -> dict:
    try:
        settings = db.reference("leaderboard/settings").get()
        if settings is None:
            settings = {
                "mode": "auto",
                "last_update": "",
                "current_prizes": {},
                "enabled": True
            }
            db.reference("leaderboard/settings").set(settings)
        return settings
    except Exception as e:
        logger.error(f"get_leaderboard_settings failed: {e}")
        return {"mode": "auto", "last_update": "", "current_prizes": {}, "enabled": True}


def set_leaderboard_settings(updates: dict):
    """Update leaderboard settings in Firebase."""
    try:
        existing = db.reference("leaderboard/settings").get() or {}
        existing.update(updates)
        db.reference("leaderboard/settings").set(existing)
    except Exception as e:
        logger.error(f"set_leaderboard_settings failed: {e}")

# ─────────────────────────────────────────────
# Leaderboard Generation
# ─────────────────────────────────────────────
def generate_real_leaderboard() -> dict:
    today_stats = get_today_stats()
    sorted_users = sorted(today_stats.items(), key=lambda x: x[1], reverse=True)[:10]
    prizes = [2.0, 1.0, 0.5, 0.5, 0.5, 0.2, 0.2, 0.2, 0.2, 0.1]
    leaderboard = {}
    for idx, (uid, count) in enumerate(sorted_users):
        if idx < len(prizes):
            leaderboard[uid] = {
                "completed": count,
                "prize": prizes[idx],
                "rank": idx + 1
            }
    return leaderboard


def generate_auto_leaderboard(increment: int = 0, previous_leaderboard: dict = None) -> dict:
    if increment and previous_leaderboard:
        leaderboard = {}
        uid_list = []
        counts = []
        for uid, data in previous_leaderboard.items():
            old_count = data.get('completed', 0)
            increment_amount = random.randint(3, 5)
            new_count = old_count + increment_amount
            uid_list.append(uid)
            counts.append(new_count)
        sorted_pairs = sorted(zip(counts, uid_list), reverse=True)
        counts = [c for c, _ in sorted_pairs]
        uid_list = [u for _, u in sorted_pairs]
        prizes = [4.0, 2.0, 1.0, 0.5, 0.5, 0.5, 0.5, 0.5, 0.5, 0.5]
        for idx, (uid, count) in enumerate(zip(uid_list, counts)):
            if idx < len(prizes):
                leaderboard[uid] = {
                    "completed": count,
                    "prize": prizes[idx],
                    "rank": idx + 1,
                    "masked_id": str(uid)[:4] + "***" + str(uid)[-2:] if len(str(uid)) > 6 else str(uid)
                }
        return leaderboard

    prizes = [4.0, 2.0, 1.0, 0.5, 0.5, 0.5, 0.5, 0.5, 0.5, 0.5]
    used_ids = set()
    uid_list = []
    while len(uid_list) < 10:
        uid = random.randint(100000000, 999999999)
        if uid not in used_ids:
            used_ids.add(uid)
            uid_list.append(str(uid))
    counts = [random.randint(10, 100) for _ in range(10)]
    counts.sort(reverse=True)
    leaderboard = {}
    for idx, (uid, count) in enumerate(zip(uid_list, counts)):
        leaderboard[uid] = {
            "completed": count,
            "prize": prizes[idx],
            "rank": idx + 1,
            "masked_id": str(uid)[:4] + "***" + str(uid)[-2:] if len(str(uid)) > 6 else str(uid)
        }
    return leaderboard


def format_leaderboard_text(leaderboard: dict, mode: str) -> str:
    if not leaderboard:
        return "📊 No data available for leaderboard yet."
    lines = [
        "🏆 **Top 10 Users Per Day**",
        "",
        "ℹ️ Results are announced every day at 01:00 (Helsinki).",
        "Leaders receive real money to their balance!",
        "",
        "🔄 Statistics update every 1PM ,6+ GTM.",
        "",
        "👤 ID          | ✅ Completed | 💰 Prize",
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
    ]
    for rank, (uid, data) in enumerate(sorted(leaderboard.items(), key=lambda x: x[1]['rank']), 1):
        completed = data.get('completed', 0)
        prize = data.get('prize', 0)
        if len(uid) > 6:
            masked = uid[:4] + "***" + uid[-2:]
        else:
            masked = uid
        if mode == "auto":
            if completed >= 70:
                completion_text = f"{completed} executions"
            elif completed >= 50:
                completion_text = f"{completed}+ executions"
            else:
                completion_text = f"{completed} executions"
        else:
            completion_text = f"{completed}"
        lines.append(f"{rank}. {masked} — {completion_text} | ${prize}")
    return "\n".join(lines)

# ─────────────────────────────────────────────
# Firebase Helpers
# ─────────────────────────────────────────────
def get_user(user_id: int) -> dict:
    try:
        ref = db.reference(f"users/{user_id}")
        data = ref.get()
        if data is None:
            data = {
                "balance": 0.0,
                "approved": 0,
                "in_review": 0,
                "total_submitted": 0,
            }
            ref.set(data)
        return data
    except Exception as e:
        logger.error(f"get_user({user_id}) failed: {e}")
        return {"balance": 0.0, "approved": 0, "in_review": 0, "total_submitted": 0}


def update_user(user_id: int, updates: dict):
    try:
        db.reference(f"users/{user_id}").update(updates)
    except Exception as e:
        logger.error(f"update_user({user_id}) failed: {e}")


def add_submission(user_id: int, tg_username: str, username: str, password: str, key: str):
    try:
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        sub_ref = db.reference(f"submissions/{user_id}").push()
        sub_ref.set({
            "username": username,
            "password": password,
            "key": key,
            "tg_username": tg_username,
            "user_id": str(user_id),
            "datetime": now,
            "status": "pending",
        })
        user = get_user(user_id)
        update_user(user_id, {
            "in_review": user.get("in_review", 0) + 1,
            "total_submitted": user.get("total_submitted", 0) + 1,
        })
        _rebuild_xlsx(user_id)
    except Exception as e:
        logger.error(f"add_submission({user_id}) failed: {e}")


def get_submissions(user_id: int) -> list:
    try:
        data = db.reference(f"submissions/{user_id}").get()
        if not data:
            return []
        return [
            {"id": k, **v} for k, v in data.items()
            if v.get("status", "pending") == "pending"
        ]
    except Exception as e:
        logger.error(f"get_submissions({user_id}) failed: {e}")
        return []


def remove_submissions(user_id: int):
    try:
        count = len(get_submissions(user_id))
        db.reference(f"submissions/{user_id}").delete()
        try:
            db.reference(f"xlsx_cache/{user_id}").delete()
        except Exception:
            pass
        user = get_user(user_id)
        new_review = max(0, user.get("in_review", 0) - count)
        update_user(user_id, {"in_review": new_review})
    except Exception as e:
        logger.error(f"remove_submissions({user_id}) failed: {e}")


def create_withdrawal(user_id: int, tg_username: str, amount: float, wallet: str) -> str:
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    ref = db.reference(f"withdrawals/{user_id}").push()
    fee = 0.025
    receive = round(amount - fee, 4)
    ref.set({
        "tg_username": tg_username,
        "amount": amount,
        "fee": fee,
        "receive": receive,
        "wallet": wallet,
        "status": "pending",
        "datetime": now,
    })
    return ref.key


def get_withdrawal(user_id: int, w_id: str):
    try:
        return db.reference(f"withdrawals/{user_id}/{w_id}").get()
    except Exception as e:
        logger.error(f"get_withdrawal({user_id}, {w_id}) failed: {e}")
        return None


def update_withdrawal(user_id: int, w_id: str, updates: dict):
    try:
        db.reference(f"withdrawals/{user_id}/{w_id}").update(updates)
    except Exception as e:
        logger.error(f"update_withdrawal({user_id}, {w_id}) failed: {e}")


def get_all_users() -> dict:
    try:
        data = db.reference("users").get()
        return data or {}
    except Exception as e:
        logger.error(f"get_all_users() failed: {e}")
        return {}


def get_all_submissions() -> dict:
    try:
        data = db.reference("submissions").get()
        return data or {}
    except Exception as e:
        logger.error(f"get_all_submissions() failed: {e}")
        return {}

# ─────────────────────────────────────────────
# Task Price Settings
# ─────────────────────────────────────────────
def get_task_price() -> float:
    try:
        price = db.reference("settings/task_price").get()
        if price is None:
            price = 0.0330
            db.reference("settings/task_price").set(price)
        return float(price)
    except Exception as e:
        logger.error(f"get_task_price failed: {e}")
        return 0.0330


def set_task_price(price: float):
    try:
        db.reference("settings/task_price").set(round(price, 4))
    except Exception as e:
        logger.error(f"set_task_price failed: {e}")


def get_task_settings() -> dict:
    try:
        raw_6h = db.reference("settings/tasks/task_6h_enabled").get()
        raw_1h = db.reference("settings/tasks/task_1h_enabled").get()
        last_updated = db.reference("settings/tasks/last_updated").get() or ""

        def parse_bool(val, default=True):
            if val is None:
                return default
            if isinstance(val, bool):
                return val
            if isinstance(val, str):
                return val.lower() != "false"
            return bool(val)

        return {
            "task_6h_enabled": parse_bool(raw_6h, True),
            "task_1h_enabled": parse_bool(raw_1h, True),
            "last_updated": last_updated,
        }
    except Exception as e:
        logger.error(f"get_task_settings failed: {e}")
        return {"task_6h_enabled": True, "task_1h_enabled": True, "last_updated": ""}


def set_task_settings(updates: dict):
    try:
        existing = db.reference("settings/tasks").get() or {}
        existing.update(updates)
        existing["last_updated"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        db.reference("settings/tasks").set(existing)
    except Exception as e:
        logger.error(f"set_task_settings failed: {e}")

# ─────────────────────────────────────────────
# XLSX Helpers
# ─────────────────────────────────────────────
XLSX_DIR = tempfile.gettempdir()


def _xlsx_path(user_id: int) -> str:
    return os.path.join(XLSX_DIR, f"submissions_{user_id}.xlsx")


def _rebuild_xlsx(user_id: int):
    try:
        subs = get_submissions(user_id)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Submissions"
        ws.append(["Username", "Password", "2FA Key", "TG Username", "User ID", "DateTime"])
        for row in subs:
            ws.append([
                row.get("username", ""),
                row.get("password", ""),
                row.get("key", ""),
                row.get("tg_username", ""),
                row.get("user_id", str(user_id)),
                row.get("datetime", ""),
            ])
        path = _xlsx_path(user_id)
        wb.save(path)
        logger.info(f"XLSX rebuilt for user {user_id} → {path}")
    except Exception as e:
        logger.error(f"_rebuild_xlsx({user_id}) failed: {e}")


def build_xlsx_bytes(user_id: int) -> bytes:
    path = _xlsx_path(user_id)
    if not os.path.exists(path):
        _rebuild_xlsx(user_id)
    try:
        with open(path, "rb") as f:
            return f.read()
    except Exception as e:
        logger.error(f"build_xlsx_bytes({user_id}) failed: {e}")
        return b""

# ─────────────────────────────────────────────
# Utility Functions
# ─────────────────────────────────────────────
UNCOMMON_ADJECTIVES = [
    "lunar", "crimson", "azure", "phantom", "mystic", "neon", "velvet",
    "cobalt", "sable", "ivory", "scarlet", "gilded", "obsidian", "spectral",
    "vivid", "prism", "hollow", "ancient", "frosted", "runic",
]
UNCOMMON_NOUNS = [
    "wraith", "oracle", "cipher", "nexus", "specter", "herald", "vortex",
    "relic", "sigil", "golem", "comet", "prism", "mirage", "scion",
    "phantom", "herald", "bastion", "ember", "zenith", "epoch",
]

_HEX_CHARS = "0123456789abcdef"


def generate_username() -> str:
    adj = random.choice(UNCOMMON_ADJECTIVES)
    noun = random.choice(UNCOMMON_NOUNS)
    num = random.randint(100, 9999)
    return f"{adj}_{noun}_{num}"


def generate_tx_hash() -> str:
    return "0x" + "".join(random.choices(_HEX_CHARS, k=64))


_BASE32_RE = re.compile(r"^[A-Z2-7]{16,32}$")


def validate_2fa_key(raw: str) -> tuple:
    original = raw.strip()
    cleaned = original.replace(" ", "").upper()
    if len(cleaned) not in (16, 32):
        return None, None, (
            f"❌ Invalid 2FA key length. Must be 16 or 32 characters "
            f"(you entered {len(cleaned)} characters after removing spaces). Try again:"
        )
    if not _BASE32_RE.match(cleaned):
        return None, None, (
            "❌ Invalid 2FA key format. Only letters A-Z and digits 2-7 are allowed. Try again:"
        )
    return cleaned, original, None


_BEP20_RE = re.compile(r"^0x[0-9a-fA-F]{40}$")


def is_valid_bep20(address: str) -> bool:
    return bool(_BEP20_RE.match(address))


# ─────────────────────────────────────────────
# TOTP Generator (fixed: hmac.new → hmac.new correct usage)
# ─────────────────────────────────────────────
def generate_totp(secret_key: str, interval: int = 30) -> str:
    """Generate 6-digit TOTP code from Base32 secret."""
    key_bytes = base64.b32decode(secret_key, casefold=True)
    current_time = int(time.time())
    time_step = current_time // interval
    time_bytes = struct.pack(">Q", time_step)
    # FIX: correct hmac usage — hmac.new(key, msg, digestmod)
    hmac_hash = hmac.new(key_bytes, time_bytes, hashlib.sha1).digest()
    offset = hmac_hash[-1] & 0x0F
    code_bytes = hmac_hash[offset:offset + 4]
    otp = struct.unpack(">I", code_bytes)[0] & 0x7FFFFFFF
    otp = otp % 1000000
    return f"{otp:06d}"

# ─────────────────────────────────────────────
# Keyboards
# ─────────────────────────────────────────────
HOME_KEYBOARD = ReplyKeyboardMarkup(
    [
        ["💰 Balance", "📋 Tasks"],
        ["📥 Withdraw", "🫟 I'm New User"],
        ["🎁 TOP"],
        ["👥 Referrals", "💝 Support"],
    ],
    resize_keyboard=True,
)

TASK_MENU_KEYBOARD = ReplyKeyboardMarkup(
    [["Inst 2FA - 0.0330$"], ["💝 Inst 2FA - 0.220$"], ["🍪 Facebook Cookie - 0.0350$"], ["🔙 Back"]],
    resize_keyboard=True,
)
TASK_START_KEYBOARD = ReplyKeyboardMarkup(
    [["✨ Start"], ["Cancel ❌"]],
    resize_keyboard=True,
)

WITHDRAW_MENU_KEYBOARD = ReplyKeyboardMarkup(
    [["USDT-BEP20"], ["BKASH - BDT"], ["🔙 Back"]],
    resize_keyboard=True,
)

BACK_KEYBOARD = ReplyKeyboardMarkup(
    [["🔙 Back"]],
    resize_keyboard=True,
)

# ─────────────────────────────────────────────
# /start
# ─────────────────────────────────────────────
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    user_id = user.id

    bot_state = get_bot_state()
    if not bot_state.get("enabled", True):
        await update.message.reply_text(
            "🔴 **Bot is Currently Offline** 🔴\n\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            "🤖 **Status:** `MAINTENANCE MODE`\n"
            f"📅 **Since:** `{bot_state.get('last_updated', 'Unknown')}`\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "⚠️ The bot is temporarily disabled by the administrator.\n\n"
            "💡 **Please try again later.**\n\n"
            "📞 For urgent inquiries, contact support:\n"
            "   👤 @Saafi_Rhman\n"
            "   👤 @its_muin",
            parse_mode=None
        )
        return HOME

    joined = await check_user_joined(user_id, context)

    if not joined and REQUIRED_CHANNEL:
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("📲 𝙹𝚘𝚒𝚗 𝙲𝚑𝚊𝚗𝚗𝚎𝚕", url=CHANNEL_LINK)],
            [InlineKeyboardButton("✔️ 𝘾𝙝𝙚𝙘𝙠", callback_data="check_join")]
        ])
        await update.message.reply_text(
            f"🔒 **Access Restricted**\n\n"
            f"👋 Welcome {user.first_name or 'User'}!\n\n"
            f"⚠️ **You must join our channel to use this bot!**\n\n"
            f"📢 Click the button below to join and then press **Check**.\n\n"
            f"✅ After joining, you'll get full access to the bot.",
            parse_mode=None,
            reply_markup=keyboard
        )
        return HOME

    return await _start_bot(update, context)


async def _start_bot(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    user_id = user.id

    if context.args and len(context.args) > 0:
        referrer_id_str = context.args[0]
        try:
            referrer_id = int(referrer_id_str)
            if referrer_id != user_id:
                user_ref_data = get_referral_data(user_id)
                if user_ref_data.get("referred_by") is None:
                    add_referral(user_id, referrer_id)
                    await update.message.reply_text(
                        "🎉 **You've been referred!** 🎉\n\n"
                        "You received a warm welcome from an existing user!\n"
                        "Complete tasks and earn money! 💰",
                        parse_mode=None
                    )
        except ValueError:
            pass

    try:
        get_user(user.id)
        get_referral_data(user.id)
    except Exception as e:
        logger.error(f"cmd_start get_user failed: {e}")

    full_name = user.full_name or user.first_name or "User"
    username = f"@{user.username}" if user.username else "No username"
    mention = f"[{full_name}](tg://user?id={user_id})"

    admin_notification = (
        f"🆕 **New User Started the Bot!**\n\n"
        f"👤 **Name:** {mention}\n"
        f"🆔 **User ID:** `{user_id}`\n"
        f"📱 **Username:** {username}\n"
        f"📅 **Time:** {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}\n\n"
        f"💡 Total users so far: `{len(get_all_users())}`"
    )

    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_message(
                chat_id=admin_id,
                text=admin_notification,
                parse_mode=None
            )
        except Exception as e:
            logger.warning(f"Could not notify admin {admin_id}: {e}")

    await update.message.reply_text(
        f"🥰 স্বাগতম, {full_name}!\n"
        "💎 কাজ শুরু করতে নিচের অপশনগুলো ব্যবহার করুন 🔽",
        reply_markup=HOME_KEYBOARD,
    )
    context.user_data.clear()
    return HOME

# ─────────────────────────────────────────────
# Channel Join Check
# ─────────────────────────────────────────────
async def check_user_joined(user_id: int, context: ContextTypes.DEFAULT_TYPE) -> bool:
    if not REQUIRED_CHANNEL:
        return True
    try:
        chat_member = await context.bot.get_chat_member(chat_id=REQUIRED_CHANNEL, user_id=user_id)
        return chat_member.status in ["member", "administrator", "creator"]
    except Exception as e:
        logger.error(f"Check join failed for {user_id}: {e}")
        return False


async def callback_check_join(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = query.from_user.id
    await query.answer()
    joined = await check_user_joined(user_id, context)
    if joined:
        await query.edit_message_text(
            "✅ **Verification Successful!**\n\n"
            "Thank you for joining our channel!\n"
            "Now you can use the bot. 🎉"
        )
        user = query.from_user
        user_id = user.id
        full_name = user.full_name or user.first_name or "User"
        username = f"@{user.username}" if user.username else "No username"
        mention = f"[{full_name}](tg://user?id={user_id})"
        try:
            get_user(user_id)
            get_referral_data(user_id)
        except Exception as e:
            logger.error(f"callback_check_join get_user failed: {e}")
        admin_notification = (
            f"🆕 **New User Started the Bot!**\n\n"
            f"👤 **Name:** {mention}\n"
            f"🆔 **User ID:** `{user_id}`\n"
            f"📱 **Username:** {username}\n"
            f"📅 **Time:** {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}\n\n"
            f"💡 Total users so far: `{len(get_all_users())}`"
        )
        for admin_id in ADMIN_IDS:
            try:
                await context.bot.send_message(
                    chat_id=admin_id,
                    text=admin_notification,
                    parse_mode=None
                )
            except Exception as e:
                logger.warning(f"Could not notify admin {admin_id}: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=f"🥰 স্বাগতম, {full_name}!\n"
                 "💎 কাজ শুরু করতে নিচের অপশনগুলো ব্যবহার করুন 🔽",
            reply_markup=HOME_KEYBOARD,
        )
        context.user_data.clear()
    else:
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("📲 𝙹𝚘𝚒𝚗 𝙲𝚑𝚊𝚗𝚗𝚎𝚕", url=CHANNEL_LINK)],
            [InlineKeyboardButton("✔️ 𝘾𝙝𝙚𝙘𝙠 𝘼𝙜𝙖𝙞𝙣", callback_data="check_join")]
        ])
        try:
            await query.edit_message_text(
                "❌ **Not Joined Yet!**\n\n"
                "You haven't joined our channel.\n\n"
                "📢 Please click the button below to join, then press **Check Again**.\n\n"
                "⚠️ Without joining, you cannot use this bot.",
                reply_markup=keyboard
            )
        except Exception as e:
            if "Message is not modified" not in str(e):
                raise

# ─────────────────────────────────────────────
# Bot Enable Check
# ─────────────────────────────────────────────
async def check_bot_enabled(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    bot_state = get_bot_state()
    if not bot_state.get("enabled", True):
        if update.effective_user and update.effective_user.id in ADMIN_IDS:
            return True
        if update.message:
            await update.message.reply_text(
                "🔴 **Bot is Currently Offline** 🔴\n\n"
                "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                "🤖 **Status:** `MAINTENANCE MODE`\n"
                f"📅 **Since:** `{bot_state.get('last_updated', 'Unknown')}`\n"
                "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                "⚠️ The bot is temporarily disabled by the administrator.\n\n"
                "💡 **Please try again later.**\n\n"
                "📞 For urgent inquiries, contact support:\n"
                "   👤 @Saafi_Rhman\n"
                "   👤 @its_muin",
                parse_mode=None
            )
        return False
    return True

# ─────────────────────────────────────────────
# HOME Handlers
# ─────────────────────────────────────────────
async def handle_balance(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    try:
        data = get_user(user_id)
        balance = data.get("balance", 0.0)
        approved = data.get("approved", 0)
        in_review = data.get("in_review", 0)
    except Exception as e:
        logger.error(f"handle_balance failed: {e}")
        await update.message.reply_text("⚠️ Could not fetch balance. Try again.", reply_markup=HOME_KEYBOARD)
        return HOME

    await update.message.reply_text(
        f"Your Balance 💰\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💵 Balance: ${balance:.4f}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"✅ Total approved: {approved}\n"
        f"⏳ In Review: {in_review}",
        reply_markup=HOME_KEYBOARD,
    )
    return HOME


async def handle_support(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        "✅ Any Problem? 24/7 Support:\n"
        "• Account related issue\n"
        "• Technical support\n"
        "• Rejected problem\n"
        "• Any query\n\n"
        "📌 Admin Contact:\n"
        "👉 @Saafi_Rhman\n"
        "👉 @its_muin",
        reply_markup=HOME_KEYBOARD,
    )
    return HOME


async def handle_new_user(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    get_user(user_id)
    try:
        db.reference("all_users").child(str(user_id)).set({
            "user_id": user_id,
            "username": update.effective_user.username,
            "joined_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        })
    except Exception as e:
        logger.error(f"Failed to add user to all_users: {e}")
    await update.message.reply_text(
        "🎁 2𝙁𝘼 𝙈𝙀𝙏𝙃𝙊𝘿?!\n\n"
        "𝙇𝙞𝙣𝙠 : [t.me](https://t.me/c/4297868120/51\n\n)"
        "💝 𝘾𝙤𝙤𝙠𝙞𝙚 𝙈𝙚𝙩𝙝𝙤𝙙?\n\n"
        "𝙇𝙞𝙣𝙠 : [t.me](https://t.me/c/4297868120/43\n\n)"
        "🛑 𝘼𝙡𝙨𝙤 𝙧𝙚𝙢𝙚𝙢𝙗𝙚𝙧 𝙒𝙚 𝙙𝙤𝙣'𝙩 𝙖𝙙𝙙𝙚𝙙 𝙘𝙤𝙤𝙠𝙞𝙚 𝙩𝙖𝙨𝙠 𝙞𝙣 𝙤𝙪𝙧 𝙗𝙤𝙩...",
        reply_markup=HOME_KEYBOARD,
        disable_web_page_preview=True,
    )
    return HOME


async def handle_leaderboard(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    settings = get_leaderboard_settings()
    if not settings.get("enabled", True):
        await update.message.reply_text(
            "🔴 **Leaderboard is currently OFF**\n\n"
            "📊 Leaderboard has been disabled by admin.\n\n"
            "💡 Please check back later."
        )
        return HOME
    mode = settings.get("mode", "auto")
    leaderboard = get_leaderboard_data()
    if not leaderboard:
        await update.message.reply_text(
            "📊 Leaderboard is being generated...\n"
            "Please check back later or contact admin."
        )
        return HOME
    text = format_leaderboard_text(leaderboard, mode)
    await update.message.reply_text(text, parse_mode=None)
    return HOME


async def handle_referrals(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    stats = get_referral_stats(user_id)
    bot_username = context.bot.username
    referral_link = f"[t.me](https://t.me/{bot_username}?start={user_id})"
    total_earned = stats.get("total_earned", 0.0)
    task_price = get_task_price()
    referral_percentage = 8

    referral_text = (
        f"👥 **Referral Program**\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📊 **Your Stats**\n"
        f"├ Total referrals: `{stats['total']}`\n"
        f"└ New in last 24h: `{stats['new_last_24h']}`\n\n"
        f"💰 **Total earned from referrals:** `${total_earned:.6f}`\n\n"
        f"🎁 **How it works:**\n"
        f"• When your referral completes a task,\n"
        f"• You get **{referral_percentage}%** of their reward!\n"
        f"• Current task reward: `${task_price:.4f}`\n"
        f"• You get: `${task_price * 0.08:.6f}` per approved task\n\n"
        f"🔗 **Your referral link:**\n"
        f"`{referral_link}`\n\n"
        f"💡 **Tip:** Share your link with friends and earn passive income!"
    )
    await update.message.reply_text(referral_text, parse_mode=None)
    return HOME

# ─────────────────────────────────────────────
# Submission Approval
# ─────────────────────────────────────────────
def approve_submission(user_id: int, submission_id: str, task_price: float):
    try:
        sub_ref = db.reference(f"submissions/{user_id}/{submission_id}")
        submission = sub_ref.get()
        if not submission:
            return False
        user_data = get_user(user_id)
        new_approved = user_data.get("approved", 0) + 1
        new_balance = round(user_data.get("balance", 0.0) + task_price, 4)
        new_in_review = max(0, user_data.get("in_review", 0) - 1)
        update_user(user_id, {
            "approved": new_approved,
            "balance": new_balance,
            "in_review": new_in_review
        })
        referral_data = get_referral_data(user_id)
        referrer_id = referral_data.get("referred_by")
        if referrer_id:
            referral_reward = round(task_price * 0.08, 6)
            if referral_reward > 0:
                referrer_data = get_user(referrer_id)
                new_referrer_balance = round(referrer_data.get("balance", 0.0) + referral_reward, 6)
                update_user(referrer_id, {"balance": new_referrer_balance})
                referrer_ref_data = get_referral_data(referrer_id)
                new_total_earned = round(referrer_ref_data.get("total_earned", 0.0) + referral_reward, 6)
                update_referral_data(referrer_id, {"total_earned": new_total_earned})
                logger.info(f"Referral reward {referral_reward} given to {referrer_id} for user {user_id}'s approval")
        # Keep the record — just mark as approved so /rcvall can retrieve it
        sub_ref.update({"status": "approved"})
        return True
    except Exception as e:
        logger.error(f"approve_submission failed: {e}")
        return False

# ─────────────────────────────────────────────
# TASKS Flow
# ─────────────────────────────────────────────
async def handle_tasks(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    task_settings = get_task_settings()
    fb_settings = get_fb_task_settings()

    keyboard = []
    if task_settings.get("task_6h_enabled", True):
        keyboard.append(["Inst 2FA - 0.0330$"])
    if task_settings.get("task_1h_enabled", True):
        keyboard.append(["💝 Inst 2FA - 0.220$"])
    if fb_settings.get("fb_task_enabled", True):
        keyboard.append(["🍪 Facebook Cookie - 0.0350$"])
    keyboard.append(["🔙 Back"])

    task_menu = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text(
        "📋 **Available Tasks:**\n\n"
        "💡 Select a task below to start earning!",
        parse_mode=None,
        reply_markup=task_menu,
    )
    return TASK_MENU


async def handle_task_2fa_info(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    task_settings = get_task_settings()
    if not task_settings.get("task_6h_enabled", True):
        await update.message.reply_text(
            "🔴 Task Temporarily Unavailable\n\n"
            "This task is currently disabled by the administrator.\n\n"
            "💡 Please try again later.\n\n"
            "📞 For inquiries, contact support:\n"
            "   👤 @axSaaFe\n"
            "   👤 @muin_007",
            reply_markup=HOME_KEYBOARD
        )
        return HOME
    await update.message.reply_text(
        "⏳ Review time: 1 minute ⏳\n\n"
        "📋 Tasks: 📱 Create Inst (2FA)\n\n"
        "📄 Description:\n"
        "In this task, you must create a new Inst acc using only a real mobile device.\n\n"
        "❗If you use your own information, your application will be REJECTED without verification.\n\n"
        "After registration:\n"
        "👉 No need to send any info\n"
        "✅ Just Send your 2fa key 🔐.\n\n"
        "⏳ Review time: 1 minute ⏳\n\n"
        f"💰 Reward: `$0.0300` per approval",
        reply_markup=TASK_START_KEYBOARD,
    )
    return TASK_2FA_INFO


async def handle_task_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    username = generate_username()
    context.user_data["task_username"] = username
    await update.message.reply_text(
        f"👤 Username:\n<code>{username}</code>\n\n"
        f"🔓 Password:\n<code>axiex@25</code>\n\n"
        "📱 Open account with above username and password.\n"
        "Then submit account with 2FA Key below 😄",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup([["Cancel ❌"]], resize_keyboard=True),
    )
    return TASK_2FA_AWAIT_KEY


async def handle_task_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.pop("task_username", None)
    await update.message.reply_text(
        "❌ Task cancelled.",
        reply_markup=HOME_KEYBOARD,
    )
    return HOME


async def handle_2fa_key(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    raw = update.message.text.strip()
    user = update.effective_user

    if raw == "Cancel ❌":
        return await handle_task_cancel(update, context)

    cleaned_key, original_key, error = validate_2fa_key(raw)
    if error:
        await update.message.reply_text(error)
        return TASK_2FA_AWAIT_KEY

    try:
        otp_code = generate_totp(cleaned_key)
    except Exception as e:
        logger.error(f"Failed to generate OTP from 2FA key: {e}")
        await update.message.reply_text(
            "⚠️ Invalid 2FA key format. Please enter a valid Base32 key.\n\n"
            "✅ Valid example: `JBSWY3DPEHPK3PXP` or `JBSWY 3DPEH PK3PX P`\n"
            "📝 Make sure the key is 16 or 32 characters long (spaces ignored).",
            parse_mode=None
        )
        return TASK_2FA_AWAIT_KEY

    context.user_data["pending_2fa_key"] = original_key if original_key else cleaned_key
    context.user_data["pending_username"] = context.user_data.get("task_username", generate_username())
    context.user_data["pending_otp_code"] = otp_code

    inline_keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 Copy 6-Digit Code", callback_data=f"copy_otp:{otp_code}")]
    ])
    reply_keyboard = ReplyKeyboardMarkup(
        [["✅ Account Registered"], ["❌ Cancel Task"]],
        resize_keyboard=True,
        one_time_keyboard=True
    )
    await update.message.reply_text(
        f"🔐 **Your 6-Digit Code:**\n"
        f"<code>{otp_code}</code>\n\n"
        f"💡 **Click the 'Copy Code' button below to copy!**\n\n"
        f"📱 **Instructions:**\n"
        f"1. Open Instagram app\n"
        f"2. Enter username: <code>{context.user_data['pending_username']}</code>\n"
        f"3. Enter password: <code>axiex@25</code>\n"
        f"4. Enter this code\n\n"
        f"✅ After login, click 'Account Registered'\n"
        f"❌ Click 'Cancel Task' to abort",
        parse_mode="HTML",
        reply_markup=reply_keyboard
    )
    await update.message.reply_text(
        "👇 **Click below to copy the code:**",
        reply_markup=inline_keyboard
    )
    return TASK_2FA_STARTED


async def _auto_approve_job(bot, user_id: int, sub_id: str, admin_msg_ids: dict) -> None:
    """Auto-approve submission if admin took no action within 1 minute."""
    await asyncio.sleep(60)

    # Check if submission still pending (not already approved/cancelled by admin)
    sub = db.reference(f"submissions/{user_id}/{sub_id}").get()
    if not sub or sub.get("status", "pending") != "pending":
        return  # Already handled by admin

    task_price = get_task_price()
    success = approve_submission(user_id, sub_id, task_price)
    if not success:
        return

    # Notify user
    try:
        await bot.send_message(
            chat_id=user_id,
            text=(
                f"✅ Your submission has been approved!\n\n"
                f"💰 +${task_price:.4f} has been added to your balance.\n\n"
                f"🎉 Thank you for completing the task!"
            ),
            parse_mode=None,
        )
    except Exception as e:
        logger.error(f"_auto_approve_job notify user failed: {e}")

    # Edit admin messages to show auto-approved
    for admin_id_str, msg_id in admin_msg_ids.items():
        try:
            await bot.edit_message_text(
                chat_id=int(admin_id_str),
                message_id=msg_id,
                text=(
                    f"✅ AUTO-APPROVED (no admin action within 1 min)\n\n"
                    f"👤 User ID: {user_id}\n"
                    f"💰 Reward: ${task_price:.4f} paid"
                ),
                parse_mode=None,
            )
        except Exception as e:
            logger.error(f"_auto_approve_job edit admin msg failed: {e}")


async def handle_account_registered(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    key = context.user_data.get("pending_2fa_key")
    ig_username = context.user_data.get("pending_username", generate_username())
    if not key:
        await update.message.reply_text(
            "⚠️ Session expired. Please start the task again.",
            reply_markup=HOME_KEYBOARD
        )
        return HOME
    password = "axiex@25"
    tg_username = f"@{user.username}" if user.username else str(user.id)
    try:
        sub_ref = db.reference(f"submissions/{user.id}").push()
        sub_id = sub_ref.key
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        sub_ref.set({
            "username": ig_username,
            "password": password,
            "key": key,
            "tg_username": tg_username,
            "user_id": str(user.id),
            "datetime": now,
            "status": "pending",
        })
        u = get_user(user.id)
        update_user(user.id, {
            "in_review": u.get("in_review", 0) + 1,
            "total_submitted": u.get("total_submitted", 0) + 1,
        })
        _rebuild_xlsx(user.id)
    except Exception as e:
        logger.error(f"handle_account_registered add_submission failed: {e}")
        await update.message.reply_text(
            "⚠️ Failed to save submission. Please try again.",
            reply_markup=HOME_KEYBOARD,
        )
        return HOME
    context.user_data.pop("pending_2fa_key", None)
    context.user_data.pop("pending_username", None)
    context.user_data.pop("task_username", None)
    context.user_data.pop("pending_otp_code", None)

    # ── Send submission to all admins with Approve/Cancel buttons ──
    task_price = get_task_price()
    submission_text = (
        f"🔔 **New Submission (Auto-approve in 1 min)**\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 User ID: `{user.id}`\n"
        f"📱 TG: {tg_username}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"🔐 Instagram Username:\n<code>{ig_username}</code>\n\n"
        f"🔑 Password:\n<code>{password}</code>\n\n"
        f"🔢 2FA Key:\n<code>{key}</code>\n\n"
        f"🕐 Submitted: {now}\n"
        f"💰 Reward: ${task_price:.4f}"
    )
    inline_kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Approve", callback_data=f"quick_approve:{user.id}:{sub_id}"),
            InlineKeyboardButton("❌ Cancel", callback_data=f"quick_cancel:{user.id}:{sub_id}"),
        ]
    ])
    admin_msg_ids = {}
    for admin_id in ADMIN_IDS:
        try:
            sent = await context.bot.send_message(
                chat_id=admin_id,
                text=submission_text,
                parse_mode="HTML",
                reply_markup=inline_kb,
            )
            admin_msg_ids[str(admin_id)] = sent.message_id
        except Exception as e:
            logger.error(f"Failed to send submission to admin {admin_id}: {e}")

    # ── Schedule auto-approve after 60 seconds (asyncio task) ──
    task = asyncio.create_task(
        _auto_approve_job(context.bot, user.id, sub_id, admin_msg_ids)
    )
    _pending_auto_approve[f"{user.id}_{sub_id}"] = task

    await update.message.reply_text(
        f"✅ **Account Successfully Registered!**\n\n"
        f"📋 Your submission has been sent for review.\n"
        f"⏳ Review time: 1 minute\n\n"
        f"💰 Upon approval, you will receive `+${task_price:.4f}`\n\n"
        f"Thank you! 🎉",
        parse_mode=None,
        reply_markup=HOME_KEYBOARD
    )
    return HOME


async def handle_2fa_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.pop("pending_2fa_key", None)
    context.user_data.pop("pending_username", None)
    context.user_data.pop("task_username", None)
    context.user_data.pop("pending_otp_code", None)
    await update.message.reply_text(
        "❌ **Task Cancelled**\n\nYou can start a new task anytime.",
        reply_markup=HOME_KEYBOARD
    )
    return HOME

# ─────────────────────────────────────────────
# 1H Task Flow
# ─────────────────────────────────────────────
async def handle_task_2fa_1h_info(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    task_settings = get_task_settings()
    if not task_settings.get("task_1h_enabled", True):
        await update.message.reply_text(
            "🔴 **Task Temporarily Unavailable** 🔴\n\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            "📋 **Task:** 2FA Instagram Account Creation\n"
            "⏳ **Review Time:** 1 Hour\n"
            "💰 **Reward:** $0.220\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "⚠️ This task is currently disabled by the administrator.\n\n"
            "💡 **Please try again later.**\n\n"
            "📞 For inquiries, contact support:\n"
            "   👤 @Saafi_Rhman\n"
            "   👤 @its_muin",
            parse_mode=None,
            reply_markup=HOME_KEYBOARD
        )
        return HOME
    context.user_data["current_task_price"] = 0.220
    context.user_data["current_task_review_time"] = "60 minutes"
    await update.message.reply_text(
        "⏳ **Review time:** 60 minutes ⏳\n\n"
        "📋 **Tasks:** 📱 Create Instagram Account (2FA Enabled)\n\n"
        "📄 **Description:**\n"
        "In this task, you must create a new Instagram account using only a real mobile device.\n\n"
        "❗ **Important:** If you use your own personal information, your application will be REJECTED without verification.\n\n"
        "**After registration:**\n"
        "👉 No need to send any personal info\n"
        "✅ Just send your 2FA backup key 🔐\n\n"
        "⏳ **Review time:** 60 minutes ⏳\n\n"
        "💰 **Reward:** `$0.220` upon approval",
        parse_mode=None,
        reply_markup=ReplyKeyboardMarkup([["Start 📑"], ["Cancel ❌"]], resize_keyboard=True),
    )
    return TASK_2FA_1H_INFO


async def handle_task_1h_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.pop("task_username", None)
    context.user_data.pop("current_task_price", None)
    context.user_data.pop("current_task_review_time", None)
    await update.message.reply_text(
        "❌ **Task Cancelled**\n\nYou can start a new task anytime from the Tasks menu.",
        reply_markup=HOME_KEYBOARD,
    )
    return HOME


async def handle_task_1h_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    username = generate_username()
    context.user_data["task_username"] = username
    await update.message.reply_text(
        f"👤 **Username:**\n<code>{username}</code>\n\n"
        f"🔓 **Password:**\n<code>axiex@25</code>\n\n"
        "📱 Open Instagram app with the above username and password.\n"
        "Then submit your 2FA backup key below 😄",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup([["Cancel ❌"]], resize_keyboard=True),
    )
    return TASK_2FA_1H_AWAIT_KEY


async def handle_2fa_key_1h(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    raw = update.message.text.strip()
    if raw == "Cancel ❌":
        return await handle_task_1h_cancel(update, context)

    cleaned_key, original_key, error = validate_2fa_key(raw)
    if error:
        await update.message.reply_text(error)
        return TASK_2FA_1H_AWAIT_KEY

    try:
        otp_code = generate_totp(cleaned_key)
    except Exception as e:
        logger.error(f"Failed to generate OTP from 2FA key: {e}")
        await update.message.reply_text(
            "⚠️ Invalid 2FA key format. Please enter a valid Base32 key.\n\n"
            "✅ Valid example: `JBSWY3DPEHPK3PXP` or `JBSWY 3DPEH PK3PX P`\n"
            "📝 Make sure the key is 16 or 32 characters long (spaces ignored)."
        )
        return TASK_2FA_1H_AWAIT_KEY

    context.user_data["pending_2fa_key"] = original_key if original_key else cleaned_key
    context.user_data["pending_username"] = context.user_data.get("task_username", generate_username())
    context.user_data["pending_otp_code"] = otp_code
    context.user_data["current_task_price"] = 0.220

    inline_keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 Copy 6-Digit Code", callback_data=f"copy_otp:{otp_code}")]
    ])
    reply_keyboard = ReplyKeyboardMarkup(
        [["✅ Account Registered"], ["❌ Cancel Task"]],
        resize_keyboard=True,
        one_time_keyboard=True
    )
    await update.message.reply_text(
        f"🔐 **Your 6-Digit Code:**\n"
        f"<code>{otp_code}</code>\n\n"
        f"💡 **Click the 'Copy Code' button below to copy!**\n\n"
        f"📱 **Instructions:**\n"
        f"1. Open Instagram app\n"
        f"2. Enter username: <code>{context.user_data['pending_username']}</code>\n"
        f"3. Enter password: <code>axiex@25</code>\n"
        f"4. Enter this code\n\n"
        f"✅ After login, click 'Account Registered'\n"
        f"❌ Click 'Cancel Task' to abort",
        parse_mode="HTML",
        reply_markup=reply_keyboard
    )
    await update.message.reply_text(
        "👇 **Click below to copy the code:**",
        reply_markup=inline_keyboard
    )
    return TASK_2FA_1H_STARTED


async def handle_account_registered_1h(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    key = context.user_data.get("pending_2fa_key")
    ig_username = context.user_data.get("pending_username", generate_username())
    task_price = context.user_data.get("current_task_price", 0.220)
    if not key:
        await update.message.reply_text(
            "⚠️ Session expired. Please start the task again.",
            reply_markup=HOME_KEYBOARD
        )
        return HOME
    password = "axiex@25"
    tg_username = f"@{user.username}" if user.username else str(user.id)
    try:
        add_submission(user.id, tg_username, ig_username, password, key)
    except Exception as e:
        logger.error(f"handle_account_registered_1h add_submission failed: {e}")
        await update.message.reply_text(
            "⚠️ Failed to save submission. Please try again.",
            reply_markup=HOME_KEYBOARD,
        )
        return HOME
    context.user_data.pop("pending_2fa_key", None)
    context.user_data.pop("pending_username", None)
    context.user_data.pop("task_username", None)
    context.user_data.pop("pending_otp_code", None)
    context.user_data.pop("current_task_price", None)
    context.user_data.pop("current_task_review_time", None)
    await update.message.reply_text(
        f"✅ **Account Successfully Registered!**\n\n"
        f"📋 Your submission has been sent for review.\n"
        f"⏳ **Review time:** 60 minutes\n\n"
        f"💰 Upon approval, you will receive `+${task_price:.4f}`\n\n"
        f"Thank you! 🎉",
        parse_mode=None,
        reply_markup=HOME_KEYBOARD
    )
    return HOME

# ─────────────────────────────────────────────
# WITHDRAW Flow
# ─────────────────────────────────────────────
async def handle_withdraw(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        "📥 **Choose withdrawal method:**\n\n"
        "💰 **USDT-BEP20** - Crypto withdrawal\n"
        "📱 **BKASH - BDT** - Direct mobile banking\n\n"
        "💡 Minimum withdrawal: $0.20 USD",
        parse_mode=None,
        reply_markup=WITHDRAW_MENU_KEYBOARD,
    )
    return WITHDRAW_MENU


async def handle_withdraw_bep20(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    try:
        data = get_user(user_id)
        balance = data.get("balance", 0.0)
    except Exception as e:
        logger.error(f"handle_withdraw_bep20 failed: {e}")
        balance = 0.0
    context.user_data["withdraw_balance"] = balance
    context.user_data["withdraw_method"] = "usdt"
    await update.message.reply_text(
        f"💵 Your current balance: ${balance:.4f}\n\n"
        "Enter the amount you want to withdraw:\n"
        "• Minimum: $0.20\n"
        "• Fee: $0.025",
        reply_markup=BACK_KEYBOARD,
    )
    return WITHDRAW_AMOUNT


async def handle_withdraw_bkash(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    try:
        data = get_user(user_id)
        balance = data.get("balance", 0.0)
    except Exception as e:
        logger.error(f"handle_withdraw_bkash failed: {e}")
        balance = 0.0
    bdt_balance = balance * USD_TO_BDT_RATE
    context.user_data["withdraw_balance"] = balance
    context.user_data["withdraw_method"] = "bkash"
    await update.message.reply_text(
        f"💵 **Your Balance:** `${balance:.4f}` USD\n"
        f"💰 **≈ {bdt_balance:.2f} BDT**\n\n"
        f"📱 **BKASH Withdrawal**\n"
        f"• Minimum withdrawal: $0.20 USD (≈ {0.20 * USD_TO_BDT_RATE:.2f} BDT)\n"
        f"• Fee: $0.025 USD (≈ {0.025 * USD_TO_BDT_RATE:.2f} BDT)\n\n"
        f"📝 Enter the amount in **USD** you want to withdraw:\n"
        f"Example: `0.50`\n\n"
        f"💡 You will receive BDT equivalent at rate 1 USD = {USD_TO_BDT_RATE} BDT",
        parse_mode=None,
        reply_markup=BACK_KEYBOARD,
    )
    return WITHDRAW_AMOUNT


async def handle_withdraw_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == "🔙 Back":
        await update.message.reply_text("🏠 Back to home.", reply_markup=HOME_KEYBOARD)
        return HOME
    try:
        amount = float(text)
    except ValueError:
        await update.message.reply_text("❌ Please enter a valid number.")
        return WITHDRAW_AMOUNT
    balance = context.user_data.get("withdraw_balance", 0.0)
    if amount < 0.20:
        await update.message.reply_text("❌ Minimum withdrawal is $0.20 USD.")
        return WITHDRAW_AMOUNT
    if amount > balance:
        await update.message.reply_text(
            f"❌ Insufficient balance. Your balance: ${balance:.4f} USD"
        )
        return WITHDRAW_AMOUNT
    context.user_data["withdraw_amount"] = amount
    method = context.user_data.get("withdraw_method", "usdt")
    if method == "bkash":
        await update.message.reply_text(
            "✅ Amount accepted!\n\n"
            "📱 Now enter your **BKASH account number** (must be a valid BKASH number):\n\n"
            "📌 Example: `01XXXXXXXXX`\n\n"
            "⚠️ Make sure the number is correct to receive payment.",
            reply_markup=BACK_KEYBOARD,
        )
    else:
        await update.message.reply_text(
            "✅ Amount accepted!\n\nNow enter your USDT (BEP-20) wallet address:",
            reply_markup=BACK_KEYBOARD,
        )
    return WITHDRAW_ADDRESS


async def handle_withdraw_address(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == "🔙 Back":
        await update.message.reply_text("🏠 Back to home.", reply_markup=HOME_KEYBOARD)
        return HOME
    method = context.user_data.get("withdraw_method", "usdt")
    amount = context.user_data.get("withdraw_amount", 0.0)
    fee = 0.025
    receive_usd = round(amount - fee, 4)
    if method == "bkash":
        bkash_pattern = re.compile(r"^01[3-9]\d{8}$")
        if not bkash_pattern.match(text):
            await update.message.reply_text(
                "❌ Invalid BKASH number.\n"
                "Must be a valid Bangladeshi mobile number.\n"
                "Example: `017XXXXXXXX` or `019XXXXXXXX`\n\n"
                "Please enter a valid BKASH number:"
            )
            return WITHDRAW_ADDRESS
        wallet = text
        receive_bdt = round(receive_usd * USD_TO_BDT_RATE, 2)
        receive_display = f"{receive_bdt:.2f} BDT (${receive_usd:.4f} USD)"
    else:
        if not is_valid_bep20(text):
            await update.message.reply_text(
                "❌ Invalid BEP-20 wallet address.\n"
                "Must start with `0x` and be exactly 42 characters long.\n"
                "Please enter a valid address:"
            )
            return WITHDRAW_ADDRESS
        wallet = text
        receive_display = f"${receive_usd:.4f}"

    user = update.effective_user
    tg_username = f"@{user.username}" if user.username else str(user.id)
    try:
        w_id = create_withdrawal(user.id, tg_username, amount, wallet)
        db.reference(f"withdrawals/{user.id}/{w_id}").update({"method": method})
    except Exception as e:
        logger.error(f"create_withdrawal failed: {e}")
        await update.message.reply_text(
            "⚠️ Failed to create withdrawal request. Please try again.",
            reply_markup=HOME_KEYBOARD,
        )
        return HOME

    context.user_data.clear()

    if method == "bkash":
        await update.message.reply_text(
            f"✅ **Withdrawal request created!**\n\n"
            f"💳 **Method:** BKASH - BDT\n"
            f"📱 **BKASH Number:** `{wallet}`\n"
            f"💵 **Debit amount:** `${amount:.4f}` USD\n"
            f"📉 **Fee:** $0.0250 USD\n"
            f"💰 **You will receive:** {receive_display}\n\n"
            f"⏳ **Processing time:** 24-48 hours\n"
            f"📞 Make sure your BKASH number is active!",
            parse_mode=None,
            reply_markup=HOME_KEYBOARD,
        )
    else:
        await update.message.reply_text(
            f"✅ **Withdrawal request created!**\n\n"
            f"💳 **Method:** USDT (BEP-20)\n"
            f"👛 **Wallet:** `{wallet}`\n"
            f"💵 **Debit amount:** `${amount:.4f}`\n"
            f"📉 **Fee:** $0.0250\n"
            f"💰 **You will receive:** {receive_display}",
            parse_mode=None,
            reply_markup=HOME_KEYBOARD,
        )

    if method == "bkash":
        admin_text = (
            f"🔔 **New Withdrawal Request**\n\n"
            f"👤 User: {tg_username}\n"
            f"🆔 Chat ID: {user.id}\n"
            f"💳 Method: BKASH - BDT\n"
            f"📱 BKASH: `{wallet}`\n"
            f"💵 Amount: ${amount:.4f} USD\n"
            f"📉 Fee: $0.0250 USD\n"
            f"💰 Receive: {receive_display}\n"
            f"🔑 Request ID: {w_id}"
        )
    else:
        admin_text = (
            f"🔔 **New Withdrawal Request**\n\n"
            f"👤 User: {tg_username}\n"
            f"🆔 Chat ID: {user.id}\n"
            f"💳 Method: USDT (BEP-20)\n"
            f"💵 Amount: ${amount:.4f}\n"
            f"📉 Fee: $0.0250\n"
            f"💰 Receive: ${receive_usd:.4f}\n"
            f"👛 Wallet: {wallet}\n"
            f"🔑 Request ID: {w_id}"
        )

    inline_kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("Approve 👍", callback_data=f"wd_approve:{user.id}:{w_id}"),
            InlineKeyboardButton("Cancel ❌", callback_data=f"wd_cancel:{user.id}:{w_id}"),
        ]
    ])
    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_message(
                chat_id=admin_id,
                text=admin_text,
                parse_mode=None,
                reply_markup=inline_kb,
            )
        except Exception as e:
            logger.warning(f"Could not notify admin {admin_id}: {e}")
    return HOME

# ─────────────────────────────────────────────
# Back → HOME
# ─────────────────────────────────────────────
async def handle_back_to_home(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    await update.message.reply_text("🏠 Back to home.", reply_markup=HOME_KEYBOARD)
    return HOME

# ─────────────────────────────────────────────
# Inline Callbacks — Withdrawal
# ─────────────────────────────────────────────
async def callback_withdrawal(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    admin_id = query.from_user.id
    if admin_id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return
    await query.answer()
    data = query.data
    try:
        action, user_id_str, w_id = data.split(":")
        user_id = int(user_id_str)
    except Exception:
        await query.edit_message_text("❌ Invalid callback data.")
        return
    try:
        wd = get_withdrawal(user_id, w_id)
    except Exception as e:
        logger.error(f"callback_withdrawal get_withdrawal failed: {e}")
        await query.edit_message_text("❌ Could not fetch withdrawal record.")
        return
    if not wd:
        await query.edit_message_text("❌ Withdrawal record not found.")
        return
    wd_ref = db.reference(f"withdrawals/{user_id}/{w_id}")

    def _atomic_status_update(current_data):
        if current_data is None:
            return None
        if current_data.get("status") != "pending":
            return None
        current_data["status"] = action
        return current_data

    try:
        result = wd_ref.transaction(_atomic_status_update)
    except Exception as e:
        logger.error(f"Firebase transaction failed: {e}")
        await query.edit_message_text("❌ Transaction error. Please try again.")
        return

    if result is None:
        existing = get_withdrawal(user_id, w_id)
        current_status = existing.get("status", "unknown") if existing else "unknown"
        await query.answer(
            f"⚠️ Request already processed ({current_status}).",
            show_alert=True,
        )
        await query.edit_message_text(
            query.message.text + f"\n\nℹ️ Already {current_status}."
        )
        return

    if action == "wd_cancel":
        update_withdrawal(user_id, w_id, {"status": "cancelled"})
        await query.edit_message_text(
            query.message.text + "\n\n❌ Request CANCELLED by admin."
        )
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text="❌ Your withdrawal request was cancelled.",
            )
        except Exception as e:
            logger.warning(f"Could not message user {user_id}: {e}")

    elif action == "wd_approve":
        wd = get_withdrawal(user_id, w_id)
        amount = wd.get("amount", 0.0)
        receive = wd.get("receive", 0.0)
        try:
            user_data = get_user(user_id)
            new_balance = max(0.0, round(user_data.get("balance", 0.0) - amount, 4))
            update_user(user_id, {"balance": new_balance})
            update_withdrawal(user_id, w_id, {"status": "approved"})
        except Exception as e:
            logger.error(f"callback_withdrawal approve update failed: {e}")
        tx_hash = generate_tx_hash()
        await query.edit_message_text(
            query.message.text + "\n\n✅ Request APPROVED by admin."
        )
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text=(
                    f"✅ Withdrawal successful!\n"
                    f"Amount: ${receive:.4f}\n\n"
                    f"Transaction:\n{tx_hash}"
                ),
            )
        except Exception as e:
            logger.warning(f"Could not message user {user_id}: {e}")

# ─────────────────────────────────────────────
# 2FA Inline Callback
# ─────────────────────────────────────────────
async def callback_2fa_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data
    if data.startswith("copy_otp:"):
        otp_code = data.split(":")[1]
        await query.answer(f"✅ Here's your code!", show_alert=False)
        await query.message.reply_text(
            f"📋 **Your 6-Digit Code:**\n"
            f"<code>{otp_code}</code>\n\n"
            f"💡 Press and hold on the code to copy it!",
            parse_mode="HTML"
        )
        return
    if data == "confirm_registered":
        user = update.effective_user
        key = context.user_data.get("pending_2fa_key")
        ig_username = context.user_data.get("pending_username", generate_username())
        if not key:
            await query.answer("❌ Session expired!", show_alert=True)
            await query.edit_message_text("⚠️ Session expired. Please start the task again.")
            return
        password = "axiex@25"
        tg_username = f"@{user.username}" if user.username else str(user.id)
        try:
            add_submission(user.id, tg_username, ig_username, password, key)
        except Exception as e:
            logger.error(f"callback_2fa_handler add_submission failed: {e}")
            await query.answer("❌ Failed to save submission!", show_alert=True)
            return
        context.user_data.pop("pending_2fa_key", None)
        context.user_data.pop("pending_username", None)
        context.user_data.pop("task_username", None)
        context.user_data.pop("pending_otp_code", None)
        await query.answer("✅ Account registered!")
        await query.edit_message_text(
            f"✅ **Account Successfully Registered!**\n\n"
            f"⏳ Review time: 1 minute\n\n"
            f"💰 Upon approval, you will receive `+${get_task_price():.4f}`\n\n"
            f"Thank you! 🎉",
            parse_mode=None
        )
        return
    if data == "cancel_2fa_task":
        context.user_data.pop("pending_2fa_key", None)
        context.user_data.pop("pending_username", None)
        context.user_data.pop("task_username", None)
        context.user_data.pop("pending_otp_code", None)
        await query.answer("❌ Task cancelled!")
        await query.edit_message_text(
            "❌ **Task Cancelled**\n\nYou can start a new task anytime.",
            parse_mode=None
        )
        return

# ─────────────────────────────────────────────
# Admin Helpers
# ─────────────────────────────────────────────
def is_admin(update: Update) -> bool:
    return update.effective_user.id in ADMIN_IDS


async def _send_long_message(bot, chat_id: int, text: str, chunk_size: int = 4000):
    lines = text.split("\n")
    chunk = ""
    for line in lines:
        if len(chunk) + len(line) + 1 > chunk_size:
            await bot.send_message(chat_id=chat_id, text=chunk)
            chunk = line + "\n"
        else:
            chunk += line + "\n"
    if chunk.strip():
        await bot.send_message(chat_id=chat_id, text=chunk)

# ─────────────────────────────────────────────
# Admin Commands
# ─────────────────────────────────────────────
async def cmd_botoff(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    args = context.args
    reason = " ".join(args) if args else "No reason provided"
    admin_id = update.effective_user.id
    try:
        set_bot_state(False, admin_id)
        await update.message.reply_text(
            f"🔴 **BOT IS NOW OFFLINE** 🔴\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🤖 **Bot Status:** `DISABLED`\n"
            f"👤 **Disabled by:** `{admin_id}`\n"
            f"📅 **Time:** `{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}`\n"
            f"📝 **Reason:** {reason}\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"⚠️ **Users will see:**\n"
            f"   '🔴 Bot is currently offline for maintenance'\n\n"
            f"✅ **To turn back ON, use:** `/boton`",
            parse_mode=None
        )
    except Exception as e:
        logger.error(f"cmd_botoff failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_boton(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    admin_id = update.effective_user.id
    try:
        set_bot_state(True, admin_id)
        await update.message.reply_text(
            f"🟢 **BOT IS NOW ONLINE** 🟢\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🤖 **Bot Status:** `ENABLED`\n"
            f"👤 **Enabled by:** `{admin_id}`\n"
            f"📅 **Time:** `{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}`\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"✅ **Bot is now fully operational!**\n"
            f"📊 Users can now use all commands.",
            parse_mode=None
        )
    except Exception as e:
        logger.error(f"cmd_boton failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_on2fa6h(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    try:
        db.reference("settings/tasks/task_6h_enabled").set("true")
        db.reference("settings/tasks/last_updated").set(
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        )
        await update.message.reply_text(
            "✅ **2FA Task (6-12h) is now ENABLED!**\n\n"
            "📋 Users can now see and start the 6-12h review task.\n"
            "💰 Reward: $0.030 per completed task"
        )
    except Exception as e:
        logger.error(f"cmd_on2fa6h failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_off2fa6h(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    try:
        db.reference("settings/tasks/task_6h_enabled").set("false")
        db.reference("settings/tasks/last_updated").set(
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        )
        verify = db.reference("settings/tasks").get()
        await update.message.reply_text(
            "🔴 **2FA Task (6-12h) is now DISABLED!**\n\n"
            "📋 Users will see a message that this task is currently unavailable.\n"
            f"✅ Verified Firebase value: `{verify}`\n"
            "💡 Use `/on2fa6h` to enable it again."
        )
    except Exception as e:
        logger.error(f"cmd_off2fa6h failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_on2fa1h(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    try:
        db.reference("settings/tasks/task_1h_enabled").set("true")
        db.reference("settings/tasks/last_updated").set(
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        )
        await update.message.reply_text(
            "✅ **2FA Task (1h) is now ENABLED!**\n\n"
            "📋 Users can now see and start the 1-hour review task.\n"
            "💰 Reward: $0.220 per completed task"
        )
    except Exception as e:
        logger.error(f"cmd_on2fa1h failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_off2fa1h(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    try:
        db.reference("settings/tasks/task_1h_enabled").set("false")
        db.reference("settings/tasks/last_updated").set(
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        )
        verify = db.reference("settings/tasks").get()
        await update.message.reply_text(
            "🔴 **2FA Task (1h) is now DISABLED!**\n\n"
            "📋 Users will see a message that this task is currently unavailable.\n"
            f"✅ Verified Firebase value: `{verify}`\n"
            "💡 Use `/on2fa1h` to enable it again."
        )
    except Exception as e:
        logger.error(f"cmd_off2fa1h failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_ldoff(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    try:
        set_leaderboard_settings({
            "enabled": False,
            "last_update": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        })
        await update.message.reply_text(
            "🔴 **Leaderboard is now OFF!**\n\n"
            "📊 Users will see: 'Leaderboard is currently disabled'\n\n"
            "💡 To turn it back ON, use: `/ldauto` or `/ldset`\n"
            "   (These commands will automatically enable it)"
        )
    except Exception as e:
        logger.error(f"cmd_ldoff failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_ldset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    status_msg = await update.message.reply_text("📊 Generating real-time leaderboard...")
    try:
        leaderboard = generate_real_leaderboard()
        set_leaderboard_data(leaderboard)
        set_leaderboard_settings({
            "mode": "real",
            "last_update": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            "enabled": True
        })
        text = format_leaderboard_text(leaderboard, "real")
        await status_msg.edit_text(text, parse_mode=None)
        await update.message.reply_text(
            f"✅ Real-time leaderboard updated!\n"
            f"📅 Last update: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        )
    except Exception as e:
        logger.error(f"cmd_ldset failed: {e}")
        await status_msg.edit_text(f"❌ Error: {e}")


async def cmd_ldauto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    status_msg = await update.message.reply_text("🎲 Generating auto leaderboard...")
    previous = get_leaderboard_data()
    try:
        settings = get_leaderboard_settings()
        is_auto_mode = settings.get("mode") == "auto"
        if is_auto_mode and previous:
            leaderboard = generate_auto_leaderboard(increment=1, previous_leaderboard=previous)
            await update.message.reply_text(
                f"📈 **Leaderboard Updated!**\n\n"
                f"✅ Each user's completed tasks increased by 3-5!\n"
                f"📅 Last update: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
            )
        else:
            leaderboard = generate_auto_leaderboard(increment=0, previous_leaderboard=None)
            await update.message.reply_text(
                f"🎲 **Auto Leaderboard Generated!**\n\n"
                f"✅ 10 fake users added to leaderboard.\n"
                f"📅 Last update: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
            )
        set_leaderboard_data(leaderboard)
        set_leaderboard_settings({
            "mode": "auto",
            "last_update": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            "enabled": True
        })
        text = format_leaderboard_text(leaderboard, "auto")
        await status_msg.edit_text(text, parse_mode=None)
    except Exception as e:
        logger.error(f"cmd_ldauto failed: {e}")
        await status_msg.edit_text(f"❌ Error: {e}")


async def cmd_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    status_msg = await update.message.reply_text("📊 Fetching statistics...")
    try:
        bot_state = get_bot_state()
        bot_status = "🟢 ONLINE" if bot_state.get("enabled", True) else "🔴 OFFLINE"
        all_users = get_all_users()
        total_users = len(all_users)
        all_subs = get_all_submissions()
        total_submissions = 0
        pending_review = 0
        for uid, subs in all_subs.items():
            if subs:
                sub_count = len(subs)
                total_submissions += sub_count
                pending_review += sub_count
        total_approved = 0
        total_balance = 0.0
        total_withdrawn = 0.0
        for uid, udata in all_users.items():
            total_approved += udata.get('approved', 0)
            total_balance += udata.get('balance', 0.0)
        try:
            all_withdrawals = db.reference("withdrawals").get()
            if all_withdrawals:
                for uid, wds in all_withdrawals.items():
                    if wds:
                        for w_id, w_data in wds.items():
                            if w_data.get('status') == 'approved':
                                total_withdrawn += w_data.get('amount', 0.0)
        except Exception as e:
            logger.warning(f"Could not fetch withdrawals for stats: {e}")
        task_price = get_task_price()
        completion_rate = (total_approved / total_submissions * 100) if total_submissions > 0 else 0
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        today_submissions = 0
        today_withdrawals = 0
        for uid, subs in all_subs.items():
            if subs:
                for sub_id, sub_data in subs.items():
                    if sub_data.get('datetime', '').startswith(today):
                        today_submissions += 1
        try:
            all_withdrawals = db.reference("withdrawals").get()
            if all_withdrawals:
                for uid, wds in all_withdrawals.items():
                    if wds:
                        for w_id, w_data in wds.items():
                            if w_data.get('datetime', '').startswith(today):
                                today_withdrawals += 1
        except Exception:
            pass
        stats_text = (
            f"📊 **BOT STATISTICS**\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🤖 **Bot Status:** {bot_status}\n"
            f"📅 **Last State Change:** `{bot_state.get('last_updated', 'Never')}`\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"👥 **USERS**\n"
            f"├ Total Users: `{total_users}`\n"
            f"└ Active Users: `{len([u for u in all_users.values() if u.get('total_submitted', 0) > 0])}`\n\n"
            f"📝 **SUBMISSIONS**\n"
            f"├ Total: `{total_submissions}`\n"
            f"├ Today: `{today_submissions}`\n"
            f"├ Pending Review: `{pending_review}`\n"
            f"└ Approved: `{total_approved}`\n\n"
            f"💰 **FINANCIALS**\n"
            f"├ Task Price: `${task_price:.4f}`\n"
            f"├ Total Balance: `${total_balance:.4f}`\n"
            f"├ Total Withdrawn: `${total_withdrawn:.4f}`\n"
            f"├ Pending Payment: `${pending_review * task_price:.4f}`\n"
            f"└ Completion Rate: `{completion_rate:.1f}%`\n\n"
            f"📈 **PERFORMANCE**\n"
            f"├ Avg per User: `{total_submissions/total_users if total_users > 0 else 0:.1f}` tasks\n"
            f"├ Today's WD: `{today_withdrawals}` requests\n"
            f"└ Potential Earnings: `${pending_review * task_price:.4f}`"
        )
        await status_msg.edit_text(stats_text, parse_mode=None)
    except Exception as e:
        logger.error(f"cmd_stats failed: {e}")
        await status_msg.edit_text(f"❌ Error fetching statistics: {e}")


async def cmd_userinfo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    if not context.args:
        await update.message.reply_text(
            "📝 **Usage:** `/userinfo {userid}`\n\n"
            "📌 **Example:** `/userinfo 123456789`",
            parse_mode=None
        )
        return
    try:
        user_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Invalid user ID. Must be a number.")
        return
    status_msg = await update.message.reply_text("🔍 Fetching user information...")
    try:
        data = get_user(user_id)
        submissions = get_submissions(user_id)
        submission_count = len(submissions)
        pending_wd = 0
        total_withdrawn = 0.0
        try:
            withdrawals = db.reference(f"withdrawals/{user_id}").get()
            if withdrawals:
                for w_id, w_data in withdrawals.items():
                    status = w_data.get("status", "")
                    amount = w_data.get("amount", 0.0)
                    if status == "pending":
                        pending_wd += 1
                    elif status == "approved":
                        total_withdrawn += amount
        except Exception as e:
            logger.warning(f"Could not fetch withdrawals for {user_id}: {e}")
        tg_username = "Not available"
        if submissions:
            for sub in submissions:
                if sub.get("tg_username"):
                    tg_username = sub.get("tg_username")
                    break
        total_earned = (data.get('approved', 0) * get_task_price())
        info_text = (
            f"👤 **User Information**\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🆔 **User ID:** `{user_id}`\n"
            f"📱 **TG Username:** `{tg_username}`\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"💰 **Current Balance:** `${data.get('balance', 0.0):.4f}`\n"
            f"💸 **Total Withdrawn:** `${total_withdrawn:.4f}`\n"
            f"💵 **Total Earned:** `${total_earned:.4f}`\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"✅ **Approved Tasks:** `{data.get('approved', 0)}`\n"
            f"⏳ **In Review:** `{data.get('in_review', 0)}`\n"
            f"📊 **Total Submitted:** `{data.get('total_submitted', 0)}`\n"
            f"📝 **Pending Submissions:** `{submission_count}`\n"
            f"⏰ **Pending Withdrawals:** `{pending_wd}`\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"💰 **Task Price:** `${get_task_price():.4f}` per task"
        )
        await status_msg.edit_text(info_text, parse_mode=None)
    except Exception as e:
        logger.error(f"cmd_userinfo failed: {e}")
        await status_msg.edit_text(f"❌ Error fetching user info: {e}")


async def cmd_stp(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    if not context.args:
        current_price = get_task_price()
        await update.message.reply_text(
            f"💰 **Current Task Price:** `${current_price:.4f}`\n\n"
            f"📝 **Usage:** `/stp 0.030`",
            parse_mode=None
        )
        return
    try:
        new_price = float(context.args[0])
        if new_price <= 0:
            await update.message.reply_text("❌ Price must be greater than 0.")
            return
        old_price = get_task_price()
        set_task_price(new_price)
        await update.message.reply_text(
            f"✅ **Task price updated!**\n\n"
            f"💰 Old price: `${old_price:.4f}`\n"
            f"💰 New price: `${new_price:.4f}`",
            parse_mode=None
        )
    except ValueError:
        await update.message.reply_text("❌ Invalid price. Example: `/stp 0.030`")
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_msg(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    if len(context.args) < 2:
        await update.message.reply_text(
            "Usage:\n/msg userid message\n\nExample:\n/msg 123456789 Hello there!"
        )
        return
    try:
        user_id = int(context.args[0])
        message = " ".join(context.args[1:])
        await context.bot.send_message(chat_id=user_id, text=message)
        await update.message.reply_text(f"✅ Message sent to user {user_id}")
    except ValueError:
        await update.message.reply_text("❌ Invalid user ID. Must be a number.")
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_cast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    if len(context.args) < 1:
        await update.message.reply_text("Usage:\n/cast message")
        return
    message = " ".join(context.args)
    status_msg = await update.message.reply_text("📢 Broadcasting message to all users...")
    try:
        all_users = get_all_users()
        if not all_users:
            await status_msg.edit_text("❌ No users found.")
            return
        success_count = 0
        fail_count = 0
        for user_id_str in all_users.keys():
            try:
                user_id = int(user_id_str)
                await context.bot.send_message(chat_id=user_id, text=message)
                success_count += 1
            except Exception as e:
                fail_count += 1
                logger.warning(f"Failed to send broadcast to {user_id_str}: {e}")
        await status_msg.edit_text(
            f"✅ Broadcast completed!\n\n"
            f"📨 Sent: {success_count} users\n"
            f"❌ Failed: {fail_count} users"
        )
    except Exception as e:
        await status_msg.edit_text(f"❌ Error: {e}")


async def cmd_add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    args = context.args
    if len(args) < 2:
        await update.message.reply_text("Usage: /add {amount} {userid}")
        return
    try:
        amount = float(args[0])
        target_id = int(args[1])
    except ValueError:
        await update.message.reply_text("❌ Invalid arguments.")
        return
    try:
        user_data = get_user(target_id)
        new_balance = round(user_data.get("balance", 0.0) + amount, 4)
        update_user(target_id, {"balance": new_balance})
        await update.message.reply_text(
            f"✅ Added ${amount:.4f} to user {target_id}.\nNew balance: ${new_balance:.4f}"
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_rm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    args = context.args
    if len(args) < 2:
        await update.message.reply_text("Usage: /rm {amount} {userid}")
        return
    try:
        amount = float(args[0])
        target_id = int(args[1])
    except ValueError:
        await update.message.reply_text("❌ Invalid arguments.")
        return
    try:
        user_data = get_user(target_id)
        new_balance = max(0.0, round(user_data.get("balance", 0.0) - amount, 4))
        update_user(target_id, {"balance": new_balance})
        await update.message.reply_text(
            f"✅ Removed ${amount:.4f} from user {target_id}.\nNew balance: ${new_balance:.4f}"
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_rmreview(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    args = context.args
    if not args:
        await update.message.reply_text("Usage: /rmreview {userid}")
        return
    try:
        target_id = int(args[0])
    except ValueError:
        await update.message.reply_text("❌ Invalid user ID.")
        return
    try:
        remove_submissions(target_id)
        update_user(target_id, {"in_review": 0})
        await update.message.reply_text(
            f"✅ All submissions removed for user {target_id}."
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_apr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    args = context.args
    if len(args) < 2:
        await update.message.reply_text("Usage: /apr {amount} {userid}")
        return
    try:
        amount = int(args[0])
        target_id = int(args[1])
    except ValueError:
        await update.message.reply_text("❌ Invalid arguments.")
        return
    try:
        user_data = get_user(target_id)
        new_approved = user_data.get("approved", 0) + amount
        new_review = max(0, user_data.get("in_review", 0) - amount)
        update_user(target_id, {"approved": new_approved, "in_review": new_review})
        await update.message.reply_text(
            f"✅ Added {amount} to approved count for user {target_id}.\n"
            f"Total approved: {new_approved}"
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_rcv(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    args = context.args
    if not args:
        await update.message.reply_text("Usage: /rcv {userid}")
        return
    try:
        target_id = int(args[0])
    except ValueError:
        await update.message.reply_text("❌ Invalid user ID.")
        return
    subs = get_submissions(target_id)
    if not subs:
        await update.message.reply_text("No submissions found for that user.")
        return
    tmp_path = None
    try:
        xlsx_data = build_xlsx_bytes(target_id)
        if not xlsx_data:
            await update.message.reply_text("❌ Could not generate XLSX.")
            return
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(xlsx_data)
            tmp_path = tmp.name
        with open(tmp_path, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=f"submissions_{target_id}.xlsx",
                caption=f"📁 Submissions for user {target_id} ({len(subs)} records)",
            )
    except Exception as e:
        logger.error(f"cmd_rcv failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


async def cmd_rcvall(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    status_msg = await update.message.reply_text("⏳ Generating XLSX for all approved submissions...")
    tmp_path = None
    try:
        all_subs = get_all_submissions()
        if not all_subs:
            await status_msg.edit_text("📭 No submissions found.")
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Approved Submissions"
        ws.append(["#", "User ID", "TG Username", "Instagram Username", "Password", "2FA Key", "Date (UTC)", "Status"])
        row_num = 1
        for user_id, submissions in all_subs.items():
            if not isinstance(submissions, dict):
                continue
            for sub_id, sub in submissions.items():
                if not isinstance(sub, dict):
                    continue
                if sub.get("status", "pending") != "approved":
                    continue
                ws.append([
                    row_num,
                    sub.get("user_id", user_id),
                    sub.get("tg_username", ""),
                    sub.get("username", ""),
                    sub.get("password", ""),
                    sub.get("key", ""),
                    sub.get("datetime", ""),
                    "approved",
                ])
                row_num += 1
        if row_num == 1:
            await status_msg.edit_text("📭 No approved submissions found yet.")
            return
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        with open(tmp_path, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename="approved_submissions.xlsx",
                caption=f"✅ Total approved submissions: {row_num - 1}",
            )
        await status_msg.delete()
    except Exception as e:
        logger.error(f"cmd_rcvall failed: {e}")
        await status_msg.edit_text(f"❌ Error: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


async def cmd_resetsub(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Delete ALL submissions (pending + approved + rejected) from Firebase."""
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    try:
        all_subs = get_all_submissions()
        total = sum(
            len(subs) if isinstance(subs, dict) else 0
            for subs in all_subs.values()
        ) if all_subs else 0

        # Delete all submissions
        db.reference("submissions").delete()
        # Also clear xlsx cache
        try:
            db.reference("xlsx_cache").delete()
        except Exception:
            pass

        # Reset in_review count for all users
        all_users = db.reference("users").get() or {}
        for uid in all_users:
            try:
                db.reference(f"users/{uid}").update({"in_review": 0})
            except Exception:
                pass

        await update.message.reply_text(
            f"🗑️ **All submissions deleted!**\n\n"
            f"📊 Total deleted: `{total}` records\n"
            f"✅ Firebase cleared successfully.",
            parse_mode=None,
        )
        logger.info(f"Admin {update.effective_user.id} reset all submissions ({total} records).")
    except Exception as e:
        logger.error(f"cmd_resetsub failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_live(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    status_msg = await update.message.reply_text("📊 Fetching active users...")
    try:
        all_users = get_all_users()
        all_subs = get_all_submissions()
        if not all_users:
            await status_msg.edit_text("❌ No users found.")
            return
        active_users = []
        for uid, udata in all_users.items():
            sub_count = len(all_subs.get(uid, {}))
            if sub_count > 0:
                active_users.append({
                    "user_id": int(uid),
                    "submissions": sub_count,
                    "balance": udata.get('balance', 0.0)
                })
        active_users.sort(key=lambda x: x['submissions'], reverse=True)
        if not active_users:
            await status_msg.edit_text("📭 No active users found.")
            return
        total_users = len(active_users)
        PAGE_SIZE = 10
        total_pages = (total_users + PAGE_SIZE - 1) // PAGE_SIZE
        context.user_data["live_users"] = active_users
        context.user_data["live_total_pages"] = total_pages
        await _send_live_page(update, context, 1)
    except Exception as e:
        logger.error(f"cmd_live failed: {e}")
        await status_msg.edit_text(f"❌ Error: {e}")


async def _send_live_page(update: Update, context: ContextTypes.DEFAULT_TYPE, page: int):
    active_users = context.user_data.get("live_users", [])
    total_pages = context.user_data.get("live_total_pages", 1)
    PAGE_SIZE = 10
    if not active_users:
        await update.message.reply_text("❌ No active users found.")
        return
    start_idx = (page - 1) * PAGE_SIZE
    end_idx = start_idx + PAGE_SIZE
    page_users = active_users[start_idx:end_idx]
    lines = [
        f"📊 **ACTIVE USERS**",
        f"📄 Page {page}/{total_pages}",
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
        ""
    ]
    for idx, user in enumerate(page_users, start=start_idx + 1):
        lines.append(
            f"`{idx}.` 🆔 **User ID:** `{user['user_id']}`\n"
            f"   📝 **Submissions:** `{user['submissions']}`\n"
            f"   💰 **Balance:** `${user['balance']:.4f}`\n"
        )
    lines.append("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    lines.append("💡 **Click a button below to copy User ID**")
    message_text = "\n".join(lines)
    keyboard = []
    for user in page_users:
        keyboard.append([
            InlineKeyboardButton(
                f"📋 Copy ID: {user['user_id']}",
                callback_data=f"live_copy_uid:{user['user_id']}"
            )
        ])
    nav_buttons = []
    if page > 1:
        nav_buttons.append(InlineKeyboardButton("◀ Previous", callback_data=f"live_page:{page - 1}"))
    if page < total_pages:
        nav_buttons.append(InlineKeyboardButton("Next ▶", callback_data=f"live_page:{page + 1}"))
    if nav_buttons:
        keyboard.append(nav_buttons)
    keyboard.append([InlineKeyboardButton("❌ Close", callback_data="live_close")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(message_text, parse_mode=None, reply_markup=reply_markup)
        await update.callback_query.answer()
    else:
        await update.message.reply_text(message_text, parse_mode=None, reply_markup=reply_markup)


async def callback_live_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    admin_id = query.from_user.id
    if admin_id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return
    data = query.data
    if data.startswith("live_page:"):
        page = int(data.split(":")[1])
        await _send_live_page(update, context, page)
        return
    if data == "live_close":
        await query.edit_message_text("✅ Live list closed.")
        await query.answer()
        return
    if data.startswith("live_copy_uid:"):
        uid = data.split(":")[1]
        await query.answer(f"✅ Copied!", show_alert=False)
        await query.message.reply_text(
            f"📋 **User ID Copied!**\n\n"
            f"<code>{uid}</code>\n\n"
            f"💡 Press and hold the code above to copy!",
            parse_mode="HTML"
        )
        return


async def cmd_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    status_msg = await update.message.reply_text("📊 Fetching user list...")
    try:
        all_users = get_all_users()
        if not all_users:
            await status_msg.edit_text("❌ No users found.")
            return
        user_ids = sorted([int(uid) for uid in all_users.keys()])
        total_users = len(user_ids)
        PAGE_SIZE = 10
        total_pages = (total_users + PAGE_SIZE - 1) // PAGE_SIZE
        context.user_data["user_list"] = user_ids
        context.user_data["total_pages"] = total_pages
        await _send_user_list_page(update, context, 1)
    except Exception as e:
        logger.error(f"cmd_list failed: {e}")
        await status_msg.edit_text(f"❌ Error: {e}")


async def _send_user_list_page(update: Update, context: ContextTypes.DEFAULT_TYPE, page: int):
    user_ids = context.user_data.get("user_list", [])
    total_pages = context.user_data.get("total_pages", 1)
    PAGE_SIZE = 10
    if not user_ids:
        await update.message.reply_text("❌ No user list found. Use /list again.")
        return
    start_idx = (page - 1) * PAGE_SIZE
    end_idx = start_idx + PAGE_SIZE
    page_users = user_ids[start_idx:end_idx]
    lines = [
        f"📋 **USER LIST** (Total: {len(user_ids)} users)",
        f"📄 Page {page}/{total_pages}",
        "━━━━━━━━━━━━━━━━━━━━━━",
        ""
    ]
    for idx, uid in enumerate(page_users, start=start_idx + 1):
        lines.append(f"{idx}. `{uid}`")
    lines.append("")
    lines.append("💡 **Click any button below to copy user ID**")
    message_text = "\n".join(lines)
    keyboard = []
    row = []
    for uid in page_users:
        row.append(InlineKeyboardButton(str(uid), callback_data=f"copy_uid:{uid}"))
        if len(row) == 3:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    nav_buttons = []
    if page > 1:
        nav_buttons.append(InlineKeyboardButton("◀ Previous", callback_data=f"list_page:{page - 1}"))
    if page < total_pages:
        nav_buttons.append(InlineKeyboardButton("Next ▶", callback_data=f"list_page:{page + 1}"))
    if nav_buttons:
        keyboard.append(nav_buttons)
    keyboard.append([InlineKeyboardButton("❌ Close", callback_data="list_close")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(
            message_text, parse_mode=None, reply_markup=reply_markup
        )
        await update.callback_query.answer()
    else:
        await update.message.reply_text(
            message_text, parse_mode=None, reply_markup=reply_markup
        )


async def callback_list_handlers(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = query.from_user.id
    if user_id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return
    data = query.data
    if data.startswith("list_page:"):
        page = int(data.split(":")[1])
        await _send_user_list_page(update, context, page)
        return
    if data == "list_close":
        await query.edit_message_text("✅ List closed.")
        await query.answer()
        return
    if data.startswith("copy_uid:"):
        uid = data.split(":")[1]
        await query.answer(f"✅ Copied: {uid}", show_alert=False)
        await query.message.reply_text(
            f"📋 User ID copied to clipboard:\n`{uid}`\n\n"
            f"💡 You can now use this ID with commands like:\n"
            f"`/add 0.5 {uid}`\n"
            f"`/userinfo {uid}`\n"
            f"`/msg {uid} Hello`",
            parse_mode=None
        )
        return


async def cmd_acts(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    args = context.args
    if not args:
        await update.message.reply_text(
            "📝 **Usage:** `/acts {userid}`\n\n"
            "📌 **Example:** `/acts 123456789`",
            parse_mode=None
        )
        return
    try:
        target_id = int(args[0])
    except ValueError:
        await update.message.reply_text("❌ Invalid user ID. Must be a number.")
        return
    context.user_data["acts_target_id"] = target_id
    submissions = get_submissions(target_id)
    if not submissions:
        await update.message.reply_text(
            f"📭 User `{target_id}` has no pending submissions.",
            parse_mode=None
        )
        return
    context.user_data["acts_pending_subs"] = submissions
    context.user_data["acts_current_index"] = 0
    context.user_data["acts_approved_count"] = 0
    context.user_data["acts_cancelled_count"] = 0
    await _send_submission_for_review(update, context, 0)
    return ADMIN_ACTS_VIEW


async def _send_submission_for_review(update: Update, context: ContextTypes.DEFAULT_TYPE, index: int):
    pending_subs = context.user_data.get("acts_pending_subs", [])
    target_id = context.user_data.get("acts_target_id")
    if not pending_subs or index >= len(pending_subs):
        approved_count = context.user_data.get("acts_approved_count", 0)
        cancelled_count = context.user_data.get("acts_cancelled_count", 0)
        summary = (
            f"✅ **Review Completed!**\n\n"
            f"👤 **User ID:** `{target_id}`\n"
            f"✅ **Approved:** `{approved_count}`\n"
            f"❌ **Cancelled:** `{cancelled_count}`"
        )
        for key in ("acts_pending_subs", "acts_target_id", "acts_current_index",
                    "acts_approved_count", "acts_cancelled_count"):
            context.user_data.pop(key, None)
        if update.callback_query:
            await update.callback_query.edit_message_text(summary, parse_mode=None)
            await update.callback_query.answer()
        else:
            await update.message.reply_text(summary, parse_mode=None)
        return
    sub = pending_subs[index]
    total = len(pending_subs)
    current = index + 1
    ig_username = sub.get('username', 'N/A')
    password = sub.get('password', 'N/A')
    twofa_key = sub.get('key', 'N/A')
    submission_text = (
        f"📋 **Submission Review**\n━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 **User ID:** `{target_id}`\n📊 **Progress:** `{current}/{total}`\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"🔐 **Instagram Username:**\n<code>{ig_username}</code>\n\n"
        f"🔑 **Password:**\n<code>{password}</code>\n\n"
        f"🔢 **2FA Key:**\n<code>{twofa_key}</code>\n\n"
        f"📱 **TG Username:** `{sub.get('tg_username', 'N/A')}`\n"
        f"🕐 **Submitted:** `{sub.get('datetime', 'N/A')}`"
    )
    keyboard = [
        [InlineKeyboardButton("📋 Copy Username", callback_data=f"acts_copy_username:{target_id}:{sub.get('id')}:{index}")],
        [InlineKeyboardButton("🔑 Copy Password", callback_data=f"acts_copy_password:{target_id}:{sub.get('id')}:{index}")],
        [InlineKeyboardButton("🔢 Copy 2FA Key", callback_data=f"acts_copy_2fa:{target_id}:{sub.get('id')}:{index}")],
        [
            InlineKeyboardButton("✅ Approve", callback_data=f"acts_approve:{target_id}:{sub.get('id')}:{index}"),
            InlineKeyboardButton("❌ Cancel", callback_data=f"acts_cancel:{target_id}:{sub.get('id')}:{index}"),
        ],
        [InlineKeyboardButton("🏠 Exit", callback_data="acts_exit")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(submission_text, parse_mode="HTML", reply_markup=reply_markup)
        await update.callback_query.answer()
    else:
        await update.message.reply_text(submission_text, parse_mode="HTML", reply_markup=reply_markup)


async def callback_acts_copy_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    admin_id = query.from_user.id
    if admin_id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return
    data = query.data
    pending_subs = context.user_data.get("acts_pending_subs", [])
    if data.startswith("acts_copy_username:"):
        parts = data.split(":")
        sub_id = parts[2]
        for sub in pending_subs:
            if sub.get('id') == sub_id:
                await query.answer("✅ Copied!", show_alert=False)
                await query.message.reply_text(
                    f"📋 **Username:**\n<code>{sub.get('username', 'N/A')}</code>\n\n💡 Press and hold to copy!",
                    parse_mode="HTML"
                )
                return
    elif data.startswith("acts_copy_password:"):
        parts = data.split(":")
        sub_id = parts[2]
        for sub in pending_subs:
            if sub.get('id') == sub_id:
                await query.answer("✅ Copied!", show_alert=False)
                await query.message.reply_text(
                    f"📋 **Password:**\n<code>{sub.get('password', 'N/A')}</code>\n\n💡 Press and hold to copy!",
                    parse_mode="HTML"
                )
                return
    elif data.startswith("acts_copy_2fa:"):
        parts = data.split(":")
        sub_id = parts[2]
        for sub in pending_subs:
            if sub.get('id') == sub_id:
                await query.answer("✅ Copied!", show_alert=False)
                await query.message.reply_text(
                    f"📋 **2FA Key:**\n<code>{sub.get('key', 'N/A')}</code>\n\n💡 Press and hold to copy!",
                    parse_mode="HTML"
                )
                return


async def callback_quick_review(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle quick Approve/Cancel from auto-sent admin submission messages."""
    query = update.callback_query
    admin_id = query.from_user.id
    if admin_id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return

    data = query.data
    parts = data.split(":")
    if len(parts) != 3:
        await query.answer("❌ Invalid data.", show_alert=True)
        return

    action, user_id_str, sub_id = parts[0], parts[1], parts[2]
    try:
        user_id = int(user_id_str)
    except ValueError:
        await query.answer("❌ Invalid user ID.", show_alert=True)
        return

    # Cancel the auto-approve task if still pending
    task_key = f"{user_id}_{sub_id}"
    pending_task = _pending_auto_approve.pop(task_key, None)
    if pending_task and not pending_task.done():
        pending_task.cancel()

    task_price = get_task_price()

    if action == "quick_approve":
        # Check if already processed
        sub = db.reference(f"submissions/{user_id}/{sub_id}").get()
        if not sub:
            await query.answer("⚠️ Already processed.", show_alert=True)
            await query.edit_message_reply_markup(reply_markup=None)
            return

        success = approve_submission(user_id, sub_id, task_price)
        if success:
            await query.edit_message_text(
                f"✅ **APPROVED** by admin\n\n"
                f"👤 User ID: `{user_id}`\n"
                f"💰 Reward: `${task_price:.4f}` paid",
                parse_mode=None,
            )
            await query.answer("✅ Approved!")
            # Notify user
            try:
                await context.bot.send_message(
                    chat_id=user_id,
                    text=(
                        f"✅ Your submission has been approved!\n\n"
                        f"💰 +${task_price:.4f} added to your balance.\n\n"
                        f"🎉 Thank you!"
                    ),
                    parse_mode=None,
                )
            except Exception as e:
                logger.error(f"quick_approve notify user failed: {e}")
        else:
            await query.answer("⚠️ Already processed.", show_alert=True)

    elif action == "quick_cancel":
        sub = db.reference(f"submissions/{user_id}/{sub_id}").get()
        if not sub:
            await query.answer("⚠️ Already processed.", show_alert=True)
            await query.edit_message_reply_markup(reply_markup=None)
            return

        try:
            db.reference(f"submissions/{user_id}/{sub_id}").update({"status": "rejected"})
            u = get_user(user_id)
            new_review = max(0, u.get("in_review", 0) - 1)
            update_user(user_id, {"in_review": new_review})
        except Exception as e:
            logger.error(f"quick_cancel update submission failed: {e}")

        await query.edit_message_text(
            f"❌ **REJECTED** by admin\n\n"
            f"👤 User ID: `{user_id}`",
            parse_mode=None,
        )
        await query.answer("❌ Rejected!")
        # Notify user
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text=(
                    f"❌ Your submission was rejected.\n\n"
                    f"Please try again with a valid account."
                ),
                parse_mode=None,
            )
        except Exception as e:
            logger.error(f"quick_cancel notify user failed: {e}")


async def callback_acts_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    admin_id = query.from_user.id
    if admin_id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return
    data = query.data
    if data == "acts_exit":
        for key in ("acts_pending_subs", "acts_target_id", "acts_current_index",
                    "acts_approved_count", "acts_cancelled_count"):
            context.user_data.pop(key, None)
        await query.edit_message_text(
            "🏠 **Review session ended.**\n\nYou can start a new review with `/acts {userid}`"
        )
        await query.answer()
        return
    if data.startswith("acts_approve:"):
        _, target_id_str, sub_id, index_str = data.split(":")
        target_id = int(target_id_str)
        index = int(index_str)
        pending_subs = context.user_data.get("acts_pending_subs", [])
        task_price = get_task_price()
        success = approve_submission(target_id, sub_id, task_price)
        if not success:
            await query.answer("❌ Failed to approve submission.", show_alert=True)
            return
        logger.info(f"Admin {admin_id} approved submission {sub_id} for user {target_id}")
        try:
            user_data = get_user(target_id)
            new_balance = user_data.get("balance", 0.0)
            await context.bot.send_message(
                chat_id=target_id,
                text=(
                    f"✅ **Report Approved!**\n\n"
                    f"💰 **Amount:** `+${task_price:.4f}`\n"
                    f"📝 **Comment:** Your Instagram account submission has been verified and approved.\n\n"
                    f"💵 **Current Balance:** `${new_balance:.4f}`\n\n"
                    f"Thank you for your hard work! 🎉"
                ),
                parse_mode=None
            )
        except Exception as e:
            logger.warning(f"Could not notify user {target_id}: {e}")
        approved_count = context.user_data.get("acts_approved_count", 0) + 1
        context.user_data["acts_approved_count"] = approved_count
        new_pending = [sub for sub in pending_subs if sub.get('id') != sub_id]
        context.user_data["acts_pending_subs"] = new_pending
        current_index = context.user_data.get("acts_current_index", 0)
        await _send_submission_for_review(update, context, current_index if new_pending else len(new_pending))
        await query.answer("✅ Submission approved!")
        return
    if data.startswith("acts_cancel:"):
        _, target_id_str, sub_id, index_str = data.split(":")
        target_id = int(target_id_str)
        index = int(index_str)
        pending_subs = context.user_data.get("acts_pending_subs", [])
        try:
            db.reference(f"submissions/{target_id}/{sub_id}").delete()
            logger.info(f"Admin {admin_id} cancelled submission {sub_id} for user {target_id}")
        except Exception as e:
            logger.error(f"Failed to delete submission {sub_id}: {e}")
            await query.answer("❌ Failed to cancel submission.", show_alert=True)
            return
        try:
            user_data = get_user(target_id)
            new_in_review = max(0, user_data.get("in_review", 0) - 1)
            update_user(target_id, {"in_review": new_in_review})
            try:
                await context.bot.send_message(
                    chat_id=target_id,
                    text=(
                        f"❌ **Report Rejected**\n\n"
                        f"📝 **Reason:** Your submitted Instagram account did not meet our quality standards.\n\n"
                        f"💡 **Tips for approval:**\n"
                        f"• Use a real mobile device for registration\n"
                        f"• Ensure 2FA is properly enabled\n"
                        f"• Submit valid 2FA backup codes\n\n"
                        f"📌 Please submit a new account following the guidelines.\n\n"
                        f"Need help? Contact support: @Saafi_Rhman / @its_muin"
                    ),
                    parse_mode=None
                )
            except Exception as e:
                logger.warning(f"Could not notify user {target_id}: {e}")
        except Exception as e:
            logger.error(f"Failed to update user {target_id}: {e}")
        cancelled_count = context.user_data.get("acts_cancelled_count", 0) + 1
        context.user_data["acts_cancelled_count"] = cancelled_count
        new_pending = [sub for sub in pending_subs if sub.get('id') != sub_id]
        context.user_data["acts_pending_subs"] = new_pending
        current_index = context.user_data.get("acts_current_index", 0)
        await _send_submission_for_review(update, context, current_index if new_pending else len(new_pending))
        await query.answer("❌ Submission cancelled/rejected!")
        return


async def cmd_checktasks(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    settings = get_task_settings()
    await update.message.reply_text(
        f"📊 **Task Settings:**\n\n"
        f"6-12h Task: `{'ENABLED' if settings.get('task_6h_enabled', True) else 'DISABLED'}`\n"
        f"1h Task: `{'ENABLED' if settings.get('task_1h_enabled', True) else 'DISABLED'}`\n\n"
        f"Last Updated: `{settings.get('last_updated', 'Never')}`"
    )


async def cmd_refreshtasks(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    await update.message.reply_text(
        "🔄 **Task menu refreshed!**\n\n"
        "Users will see updated task list when they click '📋 Tasks'."
    )


async def cmd_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    text = (
        "📋 **ADMIN COMMANDS**\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "👥 **USER MANAGEMENT**\n"
        "/add {amount} {userid} — Balance barhao\n"
        "/rm {amount} {userid} — Balance kamo\n"
        "/apr {amount} {userid} — Approved count barhao\n"
        "/rmreview {userid} — Sob submissions delete\n"
        "/userinfo {userid} — User info dekho\n"
        "/list — Sob user list (copy)\n"
        "/acts {userid} — Pending submissions review\n\n"
        "📢 **MESSAGING**\n"
        "/msg {userid} {text} — Message pathao\n"
        "/cast {text} — Broadcast koro\n\n"
        "📊 **LEADERBOARD**\n"
        "/ldset — Real leaderboard (auto enables)\n"
        "/ldauto — Auto leaderboard (auto enables)\n"
        "/ldoff — Turn leaderboard OFF\n\n"
        "🤖 **BOT CONTROL**\n"
        "/botoff — Turn bot OFF (maintenance mode)\n"
        "/boton — Turn bot back ON\n\n"
        "🔄 **TASK CONTROL**\n"
        "/on2fa6h — Enable 6-12h 2FA task ($0.030)\n"
        "/off2fa6h — Disable 6-12h 2FA task\n"
        "/on2fa1h — Enable 1h 2FA task ($0.220)\n"
        "/off2fa1h — Disable 1h 2FA task\n\n"
        "📁 **SUBMISSIONS**\n"
        "/rcv {userid} — Excel file nao\n"
        "/rcvall — All submissions Excel\n"
        "/live — Active users list\n\n"
        "⚙️ **SETTINGS**\n"
        "/stp {price} — Task price set\n"
        "/stats — Bot statistics\n"
        "/checktasks — Check task settings\n\n"
        "🍪 **FACEBOOK COOKIE TASK**\n"
        "/fbon — Enable FB Cookie task\n"
        "/fboff — Disable FB Cookie task\n"
        "/fbstp [amount] — Set FB task price\n"
        "/fblive — List users with FB submissions\n"
        "/fbrcv [userid] — Download user FB submissions (XLSX)\n"
        "/fbrcvall — Download ALL FB submissions (XLSX)\n"
        "/fbacts [userid] — Review user FB submissions\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "/cmd — Ei list"
    )
    await update.message.reply_text(text)

# ─────────────────────────────────────────────
# Fallback
# ─────────────────────────────────────────────
async def handle_unknown(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        "❓ Unknown command. Use the buttons below.",
        reply_markup=HOME_KEYBOARD,
    )
    return HOME

# ─────────────────────────────────────────────
# Error Handler
# ─────────────────────────────────────────────
async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    logger.error("Exception occurred:", exc_info=context.error)
    if isinstance(update, Update) and update.effective_message:
        try:
            await update.effective_message.reply_text(
                "⚠️ Something went wrong. Please try again."
            )
        except Exception:
            pass

# ═════════════════════════════════════════════════════════════════════════════
# FACEBOOK COOKIE TASK SYSTEM
# ═════════════════════════════════════════════════════════════════════════════

_FB_FIRSTNAMES = [
    "James","John","Robert","Michael","William","David","Richard","Joseph","Thomas","Charles",
    "Emma","Olivia","Ava","Isabella","Sophia","Mia","Charlotte","Amelia","Harper","Evelyn",
    "Daniel","Matthew","Anthony","Mark","Donald","Steven","Paul","Andrew","Kenneth","Joshua",
    "Grace","Hannah","Lily","Zoe","Ella","Nora","Aria","Chloe","Penelope","Layla",
    "Ryan","Nathan","Aaron","Adam","Brian","Eric","Tyler","Jacob","Logan","Lucas",
]

_FB_LASTNAMES = [
    "Smith","Johnson","Williams","Brown","Jones","Garcia","Miller","Davis","Wilson","Moore",
    "Taylor","Anderson","Thomas","Jackson","White","Harris","Martin","Thompson","Young","King",
    "Walker","Allen","Scott","Adams","Baker","Carter","Mitchell","Nelson","Roberts","Turner",
    "Clark","Lewis","Robinson","Lee","Hall","Perez","Wright","Hill","Green","Evans",
    "Collins","Edwards","Stewart","Morris","Rogers","Reed","Bailey","Butler","Cox","Richardson",
]


def generate_fb_firstname() -> str:
    return random.choice(_FB_FIRSTNAMES)


def generate_fb_lastname() -> str:
    return random.choice(_FB_LASTNAMES)


def get_fb_task_price() -> float:
    try:
        price = db.reference("settings/fb_task_price").get()
        if price is None:
            price = 0.035
            db.reference("settings/fb_task_price").set(price)
        return float(price)
    except Exception as e:
        logger.error(f"get_fb_task_price failed: {e}")
        return 0.035


def set_fb_task_price(price: float):
    try:
        db.reference("settings/fb_task_price").set(round(price, 4))
    except Exception as e:
        logger.error(f"set_fb_task_price failed: {e}")


def get_fb_task_settings() -> dict:
    try:
        raw = db.reference("settings/fb_tasks/fb_task_enabled").get()
        def parse_bool(val, default=True):
            if val is None:
                return default
            if isinstance(val, bool):
                return val
            if isinstance(val, str):
                return val.lower() != "false"
            return bool(val)
        return {"fb_task_enabled": parse_bool(raw, True)}
    except Exception as e:
        logger.error(f"get_fb_task_settings failed: {e}")
        return {"fb_task_enabled": True}


def set_fb_task_settings(enabled: bool):
    try:
        db.reference("settings/fb_tasks/fb_task_enabled").set("true" if enabled else "false")
    except Exception as e:
        logger.error(f"set_fb_task_settings failed: {e}")


def get_fb_user(user_id: int) -> dict:
    try:
        ref = db.reference(f"fb_users/{user_id}")
        data = ref.get()
        if data is None:
            data = {"balance": 0.0, "approved": 0, "in_review": 0, "total_submitted": 0}
            ref.set(data)
        return data
    except Exception as e:
        logger.error(f"get_fb_user({user_id}) failed: {e}")
        return {"balance": 0.0, "approved": 0, "in_review": 0, "total_submitted": 0}


def update_fb_user(user_id: int, updates: dict):
    try:
        db.reference(f"fb_users/{user_id}").update(updates)
    except Exception as e:
        logger.error(f"update_fb_user({user_id}) failed: {e}")


def add_fb_submission(user_id: int, tg_username: str, uid: str, password: str, cookies: str):
    try:
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        sub_ref = db.reference(f"fb_submissions/{user_id}").push()
        sub_ref.set({
            "uid": uid,
            "password": password,
            "cookies": cookies,
            "tg_username": tg_username,
            "user_id": str(user_id),
            "datetime": now,
        })
        fb_user = get_fb_user(user_id)
        update_fb_user(user_id, {
            "in_review": fb_user.get("in_review", 0) + 1,
            "total_submitted": fb_user.get("total_submitted", 0) + 1,
        })
        _rebuild_fb_xlsx(user_id)
        logger.info(f"FB submission saved for user {user_id}, push_id={sub_ref.key}")
    except Exception as e:
        logger.error(f"add_fb_submission({user_id}) failed: {e}")


def get_fb_submissions(user_id: int) -> list:
    try:
        data = db.reference(f"fb_submissions/{user_id}").get()
        if not data:
            return []
        return [{"id": k, **v} for k, v in data.items()]
    except Exception as e:
        logger.error(f"get_fb_submissions({user_id}) failed: {e}")
        return []


def remove_fb_submissions(user_id: int):
    try:
        count = len(get_fb_submissions(user_id))
        db.reference(f"fb_submissions/{user_id}").delete()
        try:
            db.reference(f"fb_xlsx_cache/{user_id}").delete()
        except Exception:
            pass
        fb_user = get_fb_user(user_id)
        new_review = max(0, fb_user.get("in_review", 0) - count)
        update_fb_user(user_id, {"in_review": new_review})
    except Exception as e:
        logger.error(f"remove_fb_submissions({user_id}) failed: {e}")


def approve_fb_submission(user_id: int, submission_id: str, task_price: float) -> bool:
    try:
        sub_ref = db.reference(f"fb_submissions/{user_id}/{submission_id}")
        submission = sub_ref.get()
        if not submission:
            return False
        fb_user = get_fb_user(user_id)
        update_fb_user(user_id, {
            "approved": fb_user.get("approved", 0) + 1,
            "balance": round(fb_user.get("balance", 0.0) + task_price, 4),
            "in_review": max(0, fb_user.get("in_review", 0) - 1),
        })
        sub_ref.delete()
        logger.info(f"FB submission {submission_id} approved for user {user_id}")
        return True
    except Exception as e:
        logger.error(f"approve_fb_submission failed: {e}")
        return False


def get_all_fb_submissions() -> dict:
    try:
        data = db.reference("fb_submissions").get()
        return data or {}
    except Exception as e:
        logger.error(f"get_all_fb_submissions() failed: {e}")
        return {}


def get_all_fb_users() -> dict:
    try:
        data = db.reference("fb_users").get()
        return data or {}
    except Exception as e:
        logger.error(f"get_all_fb_users() failed: {e}")
        return {}


def _fb_xlsx_path(user_id: int) -> str:
    return os.path.join(tempfile.gettempdir(), f"fb_submissions_{user_id}.xlsx")


def _rebuild_fb_xlsx(user_id: int):
    try:
        subs = get_fb_submissions(user_id)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FB Submissions"
        ws.append(["UID", "Password", "Cookies", "TG Username", "User ID", "DateTime"])
        for row in subs:
            ws.append([
                row.get("uid", ""),
                row.get("password", ""),
                row.get("cookies", ""),
                row.get("tg_username", ""),
                row.get("user_id", str(user_id)),
                row.get("datetime", ""),
            ])
        path = _fb_xlsx_path(user_id)
        wb.save(path)
        logger.info(f"FB XLSX rebuilt for user {user_id} → {path}")
    except Exception as e:
        logger.error(f"_rebuild_fb_xlsx({user_id}) failed: {e}")


def build_fb_xlsx_bytes(user_id: int) -> bytes:
    path = _fb_xlsx_path(user_id)
    if not os.path.exists(path):
        _rebuild_fb_xlsx(user_id)
    try:
        with open(path, "rb") as f:
            return f.read()
    except Exception as e:
        logger.error(f"build_fb_xlsx_bytes({user_id}) failed: {e}")
        return b""

# ─────────────────────────────────────────────
# FB User Handlers
# ─────────────────────────────────────────────
async def handle_task_fb_info(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    fb_settings = get_fb_task_settings()
    if not fb_settings.get("fb_task_enabled", True):
        await update.message.reply_text(
            "🔴 Facebook Cookie Task Temporarily Unavailable\n\n"
            "This task is currently disabled by the administrator.\n\n"
            "💡 Please try again later.\n\n"
            "📞 For inquiries, contact support:\n"
            "   👤 @Saafi_Rhman\n"
            "   👤 @its_muin",
            reply_markup=HOME_KEYBOARD,
        )
        return HOME
    price = get_fb_task_price()
    await update.message.reply_text(
        "🍪 **Facebook Cookie Task**\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "📋 **Task:** Submit a Facebook account cookie\n"
        f"💰 **Reward:** `${price:.4f}` per approved submission\n"
        "⏳ **Review time:** 1 minute\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "📄 **How it works:**\n"
        "1. We provide you a Facebook username & password\n"
        "2. Log in on a **real mobile device**\n"
        "3. Send us your **Account UID** and **Cookies** 🍪\n\n"
        "❗ Using your own personal info will cause REJECTION.\n\n"
        "✅ Click **Start 📑** to begin!",
        parse_mode=None,
        reply_markup=ReplyKeyboardMarkup(
            [["Start 📑"], ["Cancel ❌"]],
            resize_keyboard=True,
        ),
    )
    return TASK_FB_INFO


async def handle_task_fb_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    firstname = generate_fb_firstname()
    lastname = generate_fb_lastname()
    password = "axiex@25"
    context.user_data["fb_firstname"] = firstname
    context.user_data["fb_lastname"] = lastname
    context.user_data["fb_password"] = password
    inline_kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(
            f"📋 Copy First Name: {firstname}",
            callback_data=f"fb_copy:firstname:{firstname}",
        )],
        [InlineKeyboardButton(
            f"📋 Copy Last Name: {lastname}",
            callback_data=f"fb_copy:lastname:{lastname}",
        )],
        [InlineKeyboardButton(
            f"📋 Copy Password: {password}",
            callback_data=f"fb_copy:password:{password}",
        )],
    ])
    await update.message.reply_text(
        "✅ **Your Facebook Account Credentials:**\n\n"
        f"👤 **First Name:** `{firstname}`\n"
        f"👤 **Last Name:**  `{lastname}`\n"
        f"🔓 **Password:**   `{password}`\n\n"
        "📱 Log in to Facebook with these credentials on a real device.\n\n"
        "➡️ After logging in, send your **Facebook Account UID** (numeric ID).",
        parse_mode=None,
        reply_markup=ReplyKeyboardMarkup([["Cancel ❌"]], resize_keyboard=True),
    )
    await update.message.reply_text(
        "👇 **Click to copy any credential:**",
        reply_markup=inline_kb,
    )
    return TASK_FB_AWAIT_UID


async def handle_fb_await_uid(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == "Cancel ❌":
        return await handle_fb_cancel(update, context)
    if not text:
        await update.message.reply_text("⚠️ Please send your Facebook Account UID.")
        return TASK_FB_AWAIT_UID
    context.user_data["fb_uid"] = text
    logger.info(f"FB UID received from user {update.effective_user.id}: {text}")
    await update.message.reply_text(
        "✅ **Great** 👍\n\n"
        "Now send your **Account Cookies** 🍪\n\n"
        "📌 Copy and paste the full cookie string from your browser/app.\n\n"
        "⚠️ Make sure cookies are complete and valid!",
        parse_mode=None,
        reply_markup=ReplyKeyboardMarkup([["Cancel ❌"]], resize_keyboard=True),
    )
    return TASK_FB_AWAIT_COOKIES


async def handle_fb_await_cookies(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == "Cancel ❌":
        return await handle_fb_cancel(update, context)
    user = update.effective_user
    uid = context.user_data.get("fb_uid", "")
    password = context.user_data.get("fb_password", "axiex@25")
    cookies = text
    tg_username = f"@{user.username}" if user.username else str(user.id)
    price = get_fb_task_price()
    try:
        add_fb_submission(user.id, tg_username, uid, password, cookies)
    except Exception as e:
        logger.error(f"handle_fb_await_cookies add_fb_submission failed: {e}")
        await update.message.reply_text(
            "⚠️ Failed to save your submission. Please try again.",
            reply_markup=HOME_KEYBOARD,
        )
        return HOME
    context.user_data.pop("fb_uid", None)
    context.user_data.pop("fb_firstname", None)
    context.user_data.pop("fb_lastname", None)
    context.user_data.pop("fb_password", None)
    await update.message.reply_text(
        "✅ **Facebook Cookie Submitted Successfully!**\n\n"
        "📋 Your submission has been sent for review.\n"
        "⏳ Review time: 1 minute\n\n"
        f"💰 Upon approval you will receive `+${price:.4f}`\n\n"
        "Thank you! 🎉",
        parse_mode=None,
        reply_markup=HOME_KEYBOARD,
    )
    return HOME


async def handle_fb_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.pop("fb_uid", None)
    context.user_data.pop("fb_firstname", None)
    context.user_data.pop("fb_lastname", None)
    context.user_data.pop("fb_password", None)
    await update.message.reply_text(
        "❌ **Facebook Task Cancelled**\n\nYou can start a new task anytime.",
        reply_markup=HOME_KEYBOARD,
    )
    return HOME

# ─────────────────────────────────────────────
# FB Admin Commands
# ─────────────────────────────────────────────
async def cmd_fbon(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    set_fb_task_settings(True)
    await update.message.reply_text(
        "✅ **Facebook Cookie task is now ENABLED!**\n\n"
        "📋 Users can see and start the Facebook Cookie task.\n"
        f"💰 Reward: ${get_fb_task_price():.4f} per approved submission"
    )


async def cmd_fboff(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    set_fb_task_settings(False)
    await update.message.reply_text(
        "🔴 **Facebook Cookie task is now DISABLED!**\n\n"
        "📋 Users will see a message that this task is unavailable.\n"
        "💡 Use `/fbon` to enable it again."
    )


async def cmd_fbstp(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    if not context.args:
        current = get_fb_task_price()
        await update.message.reply_text(
            f"💰 **Current FB Task Price:** `${current:.4f}`\n\n"
            "📝 **Usage:** `/fbstp 0.030`",
            parse_mode=None,
        )
        return
    try:
        new_price = float(context.args[0])
        if new_price <= 0:
            await update.message.reply_text("❌ Price must be greater than 0.")
            return
        old_price = get_fb_task_price()
        set_fb_task_price(new_price)
        await update.message.reply_text(
            f"✅ **FB Task price updated!**\n\n"
            f"💰 Old price: `${old_price:.4f}`\n"
            f"💰 New price: `${new_price:.4f}`",
            parse_mode=None,
        )
    except ValueError:
        await update.message.reply_text("❌ Invalid price. Use a number e.g. `/fbstp 0.030`")
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_fblive(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    status_msg = await update.message.reply_text("📊 Fetching FB active users...")
    try:
        all_subs = get_all_fb_submissions()
        all_fb_users = get_all_fb_users()
        if not all_subs:
            await status_msg.edit_text("📭 No Facebook submissions found.")
            return
        active_users = []
        for uid, subs in all_subs.items():
            if isinstance(subs, dict) and subs:
                fb_data = all_fb_users.get(uid, {})
                active_users.append({
                    "user_id": int(uid),
                    "submissions": len(subs),
                    "balance": fb_data.get("balance", 0.0),
                })
        active_users.sort(key=lambda x: x["submissions"], reverse=True)
        if not active_users:
            await status_msg.edit_text("📭 No active FB users found.")
            return
        PAGE_SIZE = 10
        total_pages = (len(active_users) + PAGE_SIZE - 1) // PAGE_SIZE
        context.user_data["fb_live_users"] = active_users
        context.user_data["fb_live_total_pages"] = total_pages
        await _send_fb_live_page(update, context, 1)
    except Exception as e:
        logger.error(f"cmd_fblive failed: {e}")
        await status_msg.edit_text(f"❌ Error: {e}")


async def _send_fb_live_page(update: Update, context: ContextTypes.DEFAULT_TYPE, page: int):
    active_users = context.user_data.get("fb_live_users", [])
    total_pages = context.user_data.get("fb_live_total_pages", 1)
    PAGE_SIZE = 10
    if not active_users:
        await update.message.reply_text("❌ No FB active users found.")
        return
    start_idx = (page - 1) * PAGE_SIZE
    end_idx = start_idx + PAGE_SIZE
    page_users = active_users[start_idx:end_idx]
    lines = [
        "🍪 **FB ACTIVE USERS**",
        f"📄 Page {page}/{total_pages}",
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
        "",
    ]
    for idx, u in enumerate(page_users, start=start_idx + 1):
        lines.append(
            f"`{idx}.` 🆔 **User ID:** `{u['user_id']}`\n"
            f"   📝 **Submissions:** `{u['submissions']}`\n"
            f"   💰 **Balance:** `${u['balance']:.4f}`\n"
        )
    lines.append("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    lines.append("💡 **Click a button below to copy User ID**")
    message_text = "\n".join(lines)
    keyboard = []
    for u in page_users:
        keyboard.append([
            InlineKeyboardButton(
                f"📋 Copy ID: {u['user_id']}",
                callback_data=f"fb_live_copy_uid:{u['user_id']}",
            )
        ])
    nav = []
    if page > 1:
        nav.append(InlineKeyboardButton("◀ Previous", callback_data=f"fb_live_page:{page - 1}"))
    if page < total_pages:
        nav.append(InlineKeyboardButton("Next ▶", callback_data=f"fb_live_page:{page + 1}"))
    if nav:
        keyboard.append(nav)
    keyboard.append([InlineKeyboardButton("❌ Close", callback_data="fb_live_close")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(message_text, parse_mode=None, reply_markup=reply_markup)
        await update.callback_query.answer()
    else:
        await update.message.reply_text(message_text, parse_mode=None, reply_markup=reply_markup)


async def callback_fb_live_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if query.from_user.id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return
    data = query.data
    if data.startswith("fb_live_page:"):
        page = int(data.split(":")[1])
        await _send_fb_live_page(update, context, page)
        return
    if data == "fb_live_close":
        await query.edit_message_text("✅ FB live list closed.")
        await query.answer()
        return
    if data.startswith("fb_live_copy_uid:"):
        uid = data.split(":")[1]
        await query.answer("✅ Copied!", show_alert=False)
        await query.message.reply_text(
            f"📋 **FB User ID Copied!**\n\n"
            f"<code>{uid}</code>\n\n"
            "💡 Press and hold the code above to copy!\n\n"
            f"Use with: `/fbrcv {uid}` or `/fbacts {uid}`",
            parse_mode="HTML",
        )
        return


async def cmd_fbrcv(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    if not context.args:
        await update.message.reply_text("Usage: /fbrcv {userid}")
        return
    try:
        target_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Invalid user ID.")
        return
    subs = get_fb_submissions(target_id)
    if not subs:
        await update.message.reply_text(f"📭 No FB submissions for user {target_id}.")
        return
    tmp_path = None
    try:
        xlsx_data = build_fb_xlsx_bytes(target_id)
        if not xlsx_data:
            await update.message.reply_text("❌ Could not generate XLSX.")
            return
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(xlsx_data)
            tmp_path = tmp.name
        with open(tmp_path, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=f"fb_submissions_{target_id}.xlsx",
                caption=f"🍪 FB Submissions for user {target_id} ({len(subs)} records)",
            )
    except Exception as e:
        logger.error(f"cmd_fbrcv failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


async def cmd_fbrcvall(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    status_msg = await update.message.reply_text("⏳ Generating FB XLSX for all submissions...")
    tmp_path = None
    try:
        all_subs = get_all_fb_submissions()
        if not all_subs:
            await status_msg.edit_text("📭 No FB submissions found.")
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All FB Submissions"
        ws.append(["#", "UID", "Password", "Cookies", "TG Username", "User ID", "DateTime"])
        row_num = 1
        for user_id, submissions in all_subs.items():
            if not isinstance(submissions, dict):
                continue
            for sub_id, sub in submissions.items():
                if not isinstance(sub, dict):
                    continue
                ws.append([
                    row_num,
                    sub.get("uid", ""),
                    sub.get("password", ""),
                    sub.get("cookies", ""),
                    sub.get("tg_username", ""),
                    sub.get("user_id", str(user_id)),
                    sub.get("datetime", ""),
                ])
                row_num += 1
        if row_num == 1:
            await status_msg.edit_text("📭 No FB submissions found.")
            return
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        with open(tmp_path, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename="all_fb_submissions.xlsx",
                caption=f"🍪 Total FB submissions: {row_num - 1}",
            )
        await status_msg.delete()
    except Exception as e:
        logger.error(f"cmd_fbrcvall failed: {e}")
        await status_msg.edit_text(f"❌ Error: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


async def cmd_fbacts(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    if not context.args:
        await update.message.reply_text(
            "📝 **Usage:** `/fbacts {userid}`\n\n"
            "📌 **Example:** `/fbacts 123456789`",
            parse_mode=None,
        )
        return
    try:
        target_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Invalid user ID. Must be a number.")
        return
    context.user_data["fbacts_target_id"] = target_id
    submissions = get_fb_submissions(target_id)
    if not submissions:
        await update.message.reply_text(
            f"📭 User `{target_id}` has no pending FB submissions.",
            parse_mode=None,
        )
        return
    context.user_data["fbacts_pending_subs"] = submissions
    context.user_data["fbacts_current_index"] = 0
    context.user_data["fbacts_approved_count"] = 0
    context.user_data["fbacts_cancelled_count"] = 0
    await _send_fb_submission_for_review(update, context, 0)
    return ADMIN_FBACTS_VIEW


async def _send_fb_submission_for_review(update: Update, context: ContextTypes.DEFAULT_TYPE, index: int):
    pending_subs = context.user_data.get("fbacts_pending_subs", [])
    target_id = context.user_data.get("fbacts_target_id")
    if not pending_subs or index >= len(pending_subs):
        approved_count = context.user_data.get("fbacts_approved_count", 0)
        cancelled_count = context.user_data.get("fbacts_cancelled_count", 0)
        summary = (
            f"✅ **FB Review Completed!**\n\n"
            f"👤 **User ID:** `{target_id}`\n"
            f"✅ **Approved:** `{approved_count}`\n"
            f"❌ **Cancelled:** `{cancelled_count}`"
        )
        for key in ("fbacts_pending_subs", "fbacts_target_id", "fbacts_current_index",
                    "fbacts_approved_count", "fbacts_cancelled_count"):
            context.user_data.pop(key, None)
        if update.callback_query:
            await update.callback_query.edit_message_text(summary, parse_mode=None)
            await update.callback_query.answer()
        else:
            await update.message.reply_text(summary, parse_mode=None)
        return
    sub = pending_subs[index]
    total = len(pending_subs)
    current = index + 1
    fb_uid = sub.get("uid", "N/A")
    password = sub.get("password", "N/A")
    cookies = sub.get("cookies", "N/A")
    cookies_display = (cookies[:200] + "...") if len(cookies) > 200 else cookies
    submission_text = (
        f"🍪 **FB Submission Review**\n━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 **User ID:** `{target_id}`\n📊 **Progress:** `{current}/{total}`\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"🆔 **Facebook UID:**\n<code>{fb_uid}</code>\n\n"
        f"🔑 **Password:**\n<code>{password}</code>\n\n"
        f"🍪 **Cookies (preview):**\n<code>{cookies_display}</code>\n\n"
        f"📱 **TG Username:** `{sub.get('tg_username', 'N/A')}`\n"
        f"🕐 **Submitted:** `{sub.get('datetime', 'N/A')}`"
    )
    sub_id = sub.get("id")
    keyboard = [
        [InlineKeyboardButton("📋 Copy UID", callback_data=f"fbacts_copy_uid:{target_id}:{sub_id}:{index}")],
        [InlineKeyboardButton("🔑 Copy Password", callback_data=f"fbacts_copy_password:{target_id}:{sub_id}:{index}")],
        [InlineKeyboardButton("🍪 Copy Cookies", callback_data=f"fbacts_copy_cookies:{target_id}:{sub_id}:{index}")],
        [
            InlineKeyboardButton("✅ Approve", callback_data=f"fbacts_approve:{target_id}:{sub_id}:{index}"),
            InlineKeyboardButton("❌ Reject", callback_data=f"fbacts_cancel:{target_id}:{sub_id}:{index}"),
        ],
        [InlineKeyboardButton("🏠 Exit", callback_data="fbacts_exit")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(submission_text, parse_mode="HTML", reply_markup=reply_markup)
        await update.callback_query.answer()
    else:
        await update.message.reply_text(submission_text, parse_mode="HTML", reply_markup=reply_markup)


async def callback_fb_copy_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if query.from_user.id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return
    data = query.data
    pending_subs = context.user_data.get("fbacts_pending_subs", [])
    if data.startswith("fbacts_copy_uid:"):
        parts = data.split(":")
        sub_id = parts[2]
        for sub in pending_subs:
            if sub.get("id") == sub_id:
                await query.answer("✅ Copied!", show_alert=False)
                await query.message.reply_text(
                    f"📋 **Facebook UID:**\n<code>{sub.get('uid', 'N/A')}</code>\n\n💡 Press and hold to copy!",
                    parse_mode="HTML",
                )
                return
    elif data.startswith("fbacts_copy_password:"):
        parts = data.split(":")
        sub_id = parts[2]
        for sub in pending_subs:
            if sub.get("id") == sub_id:
                await query.answer("✅ Copied!", show_alert=False)
                await query.message.reply_text(
                    f"📋 **Password:**\n<code>{sub.get('password', 'N/A')}</code>\n\n💡 Press and hold to copy!",
                    parse_mode="HTML",
                )
                return
    elif data.startswith("fbacts_copy_cookies:"):
        parts = data.split(":")
        sub_id = parts[2]
        for sub in pending_subs:
            if sub.get("id") == sub_id:
                await query.answer("✅ Copied!", show_alert=False)
                cookies = sub.get("cookies", "N/A")
                await query.message.reply_text(
                    f"📋 **Cookies:**\n<code>{cookies}</code>\n\n💡 Press and hold to copy!",
                    parse_mode="HTML",
                )
                return
    await query.answer("⚠️ Submission not found in session.", show_alert=True)


async def callback_fb_acts_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    admin_id = query.from_user.id
    if admin_id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return
    data = query.data
    if data == "fbacts_exit":
        for key in ("fbacts_pending_subs", "fbacts_target_id", "fbacts_current_index",
                    "fbacts_approved_count", "fbacts_cancelled_count"):
            context.user_data.pop(key, None)
        await query.edit_message_text(
            "🏠 **FB Review session ended.**\n\nStart a new review with `/fbacts {userid}`"
        )
        await query.answer()
        return
    if data.startswith("fbacts_approve:"):
        _, target_id_str, sub_id, index_str = data.split(":")
        target_id = int(target_id_str)
        pending_subs = context.user_data.get("fbacts_pending_subs", [])
        task_price = get_fb_task_price()
        success = approve_fb_submission(target_id, sub_id, task_price)
        if not success:
            await query.answer("❌ Failed to approve.", show_alert=True)
            return
        logger.info(f"Admin {admin_id} approved FB submission {sub_id} for user {target_id}")
        try:
            fb_user_data = get_fb_user(target_id)
            new_balance = fb_user_data.get("balance", 0.0)
            await context.bot.send_message(
                chat_id=target_id,
                text=(
                    f"✅ **FB Cookie Report Approved!**\n\n"
                    f"💰 **Amount:** `+${task_price:.4f}`\n"
                    f"📝 Your Facebook Cookie submission has been verified and approved.\n\n"
                    f"💵 **Current FB Balance:** `${new_balance:.4f}`\n\n"
                    f"Thank you! 🎉"
                ),
                parse_mode=None,
            )
        except Exception as e:
            logger.warning(f"Could not notify user {target_id}: {e}")
        approved_count = context.user_data.get("fbacts_approved_count", 0) + 1
        context.user_data["fbacts_approved_count"] = approved_count
        new_pending = [s for s in pending_subs if s.get("id") != sub_id]
        context.user_data["fbacts_pending_subs"] = new_pending
        current_index = context.user_data.get("fbacts_current_index", 0)
        await _send_fb_submission_for_review(update, context, current_index if new_pending else len(new_pending))
        await query.answer("✅ FB submission approved!")
        return
    if data.startswith("fbacts_cancel:"):
        _, target_id_str, sub_id, index_str = data.split(":")
        target_id = int(target_id_str)
        pending_subs = context.user_data.get("fbacts_pending_subs", [])
        try:
            db.reference(f"fb_submissions/{target_id}/{sub_id}").delete()
            logger.info(f"Admin {admin_id} rejected FB submission {sub_id} for user {target_id}")
        except Exception as e:
            logger.error(f"Failed to delete FB submission {sub_id}: {e}")
            await query.answer("❌ Failed to reject.", show_alert=True)
            return
        try:
            fb_user_data = get_fb_user(target_id)
            new_in_review = max(0, fb_user_data.get("in_review", 0) - 1)
            update_fb_user(target_id, {"in_review": new_in_review})
            try:
                await context.bot.send_message(
                    chat_id=target_id,
                    text=(
                        "❌ **FB Cookie Report Rejected**\n\n"
                        "📝 **Reason:** Your submitted Facebook cookie did not meet our requirements.\n\n"
                        "💡 **Tips for approval:**\n"
                        "• Use a real mobile device\n"
                        "• Ensure the cookie is fresh and complete\n"
                        "• Submit the correct UID\n\n"
                        "📌 Please submit again following the guidelines.\n\n"
                        "Need help? Contact support: @Saafi_Rhman / @its_muin"
                    ),
                    parse_mode=None,
                )
            except Exception as e:
                logger.warning(f"Could not notify user {target_id}: {e}")
        except Exception as e:
            logger.error(f"Failed to update FB user {target_id}: {e}")
        cancelled_count = context.user_data.get("fbacts_cancelled_count", 0) + 1
        context.user_data["fbacts_cancelled_count"] = cancelled_count
        new_pending = [s for s in pending_subs if s.get("id") != sub_id]
        context.user_data["fbacts_pending_subs"] = new_pending
        current_index = context.user_data.get("fbacts_current_index", 0)
        await _send_fb_submission_for_review(update, context, current_index if new_pending else len(new_pending))
        await query.answer("❌ FB submission rejected!")
        return


async def callback_fb_copy_user_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data  # fb_copy:fieldname:value
    parts = data.split(":", 2)
    if len(parts) < 3:
        await query.answer("⚠️ Invalid callback.", show_alert=True)
        return
    field = parts[1]
    value = parts[2]
    label_map = {"firstname": "First Name", "lastname": "Last Name", "password": "Password"}
    label = label_map.get(field, field.capitalize())
    await query.answer("✅ See message below!", show_alert=False)
    await query.message.reply_text(
        f"📋 **{label}:**\n<code>{value}</code>\n\n💡 Press and hold to copy!",
        parse_mode="HTML",
    )

# ═════════════════════════════════════════════════════════════════════════════
# Main
# ═════════════════════════════════════════════════════════════════════════════
def main():
    init_firebase()

    app = Application.builder().token(BOT_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", cmd_start)],
        states={
            HOME: [
                MessageHandler(filters.Regex("^💰 Balance$"), handle_balance),
                MessageHandler(filters.Regex("^📋 Tasks$"), handle_tasks),
                MessageHandler(filters.Regex("^🎁 TOP"), handle_leaderboard),
                MessageHandler(filters.Regex("^👥 Referrals$"), handle_referrals),
                MessageHandler(filters.Regex("^📥 Withdraw$"), handle_withdraw),
                MessageHandler(filters.Regex("^🫟 I'm New User$"), handle_new_user),
                MessageHandler(filters.Regex("^💝 Support$"), handle_support),
            ],
            TASK_MENU: [
                # FIX: regex now matches actual keyboard labels
                MessageHandler(filters.Regex(r"^Inst 2FA - 0\.0330\$"), handle_task_2fa_info),
                MessageHandler(filters.Regex(r"^💝 Inst 2FA - 0\.220\$"), handle_task_2fa_1h_info),
                MessageHandler(filters.Regex(r"^🍪 Facebook Cookie - 0\.0350\$"), handle_task_fb_info),
                MessageHandler(filters.Regex("^🔙 Back$"), handle_back_to_home),
            ],
            TASK_FB_INFO: [
                MessageHandler(filters.Regex("^Start 📑$"), handle_task_fb_start),
                MessageHandler(filters.Regex("^Cancel ❌$"), handle_fb_cancel),
                MessageHandler(filters.Regex("^🔙 Back$"), handle_back_to_home),
            ],
            TASK_FB_AWAIT_UID: [
                MessageHandler(filters.Regex("^Cancel ❌$"), handle_fb_cancel),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_fb_await_uid),
            ],
            TASK_FB_AWAIT_COOKIES: [
                MessageHandler(filters.Regex("^Cancel ❌$"), handle_fb_cancel),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_fb_await_cookies),
            ],
            TASK_FB_STARTED: [
                MessageHandler(filters.Regex("^Cancel ❌$"), handle_fb_cancel),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_fb_await_cookies),
            ],
            TASK_2FA_INFO: [
                MessageHandler(filters.Regex("^Cancel ❌$"), handle_task_cancel),
                MessageHandler(filters.Regex("^✨ Start$"), handle_task_start),
                MessageHandler(filters.Regex("^🔙 Back$"), handle_back_to_home),
            ],
            TASK_2FA_STARTED: [
                MessageHandler(filters.Regex("^✅ Account Registered$"), handle_account_registered),
                MessageHandler(filters.Regex("^❌ Cancel Task$"), handle_2fa_cancel),
            ],
            TASK_2FA_AWAIT_KEY: [
                MessageHandler(filters.Regex("^Cancel ❌$"), handle_task_cancel),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_2fa_key),
            ],
            TASK_2FA_1H_INFO: [
                MessageHandler(filters.Regex("^Cancel ❌$"), handle_task_1h_cancel),
                MessageHandler(filters.Regex("^Start 📑$"), handle_task_1h_start),
                MessageHandler(filters.Regex("^🔙 Back$"), handle_back_to_home),
            ],
            TASK_2FA_1H_AWAIT_KEY: [
                MessageHandler(filters.Regex("^Cancel ❌$"), handle_task_1h_cancel),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_2fa_key_1h),
            ],
            TASK_2FA_1H_STARTED: [
                MessageHandler(filters.Regex("^✅ Account Registered$"), handle_account_registered_1h),
                MessageHandler(filters.Regex("^❌ Cancel Task$"), handle_task_1h_cancel),
            ],
            WITHDRAW_MENU: [
                MessageHandler(filters.Regex("^USDT-BEP20$"), handle_withdraw_bep20),
                MessageHandler(filters.Regex("^BKASH - BDT$"), handle_withdraw_bkash),
                MessageHandler(filters.Regex("^🔙 Back$"), handle_back_to_home),
            ],
            WITHDRAW_AMOUNT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_withdraw_amount),
            ],
            WITHDRAW_ADDRESS: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_withdraw_address),
            ],
        },
        fallbacks=[
            CommandHandler("start", cmd_start),
            MessageHandler(filters.TEXT & ~filters.COMMAND, handle_unknown),
        ],
        allow_reentry=True,
    )

    app.add_handler(conv)

    # Callback handlers
    app.add_handler(CallbackQueryHandler(callback_withdrawal, pattern="^wd_(approve|cancel):"))
    app.add_handler(CallbackQueryHandler(callback_live_handler, pattern="^(live_page:|live_copy_uid:|live_close)"))
    app.add_handler(CallbackQueryHandler(callback_check_join, pattern="^check_join$"))
    app.add_handler(CallbackQueryHandler(callback_acts_copy_handler, pattern="^(acts_copy_username:|acts_copy_password:|acts_copy_2fa:)"))
    app.add_handler(CallbackQueryHandler(callback_2fa_handler, pattern="^(copy_otp:|confirm_registered|cancel_2fa_task)"))
    app.add_handler(CallbackQueryHandler(callback_list_handlers, pattern="^(list_page:|copy_uid:|list_close)"))
    app.add_handler(CallbackQueryHandler(callback_quick_review, pattern="^(quick_approve:|quick_cancel:)"))
    app.add_handler(CallbackQueryHandler(callback_acts_handler, pattern="^(acts_approve:|acts_cancel:|acts_exit)"))
    # FB callbacks
    app.add_handler(CallbackQueryHandler(callback_fb_copy_user_handler, pattern="^fb_copy:"))
    app.add_handler(CallbackQueryHandler(callback_fb_live_handler, pattern="^(fb_live_page:|fb_live_copy_uid:|fb_live_close)"))
    app.add_handler(CallbackQueryHandler(callback_fb_copy_handler, pattern="^(fbacts_copy_uid:|fbacts_copy_password:|fbacts_copy_cookies:)"))
    app.add_handler(CallbackQueryHandler(callback_fb_acts_handler, pattern="^(fbacts_approve:|fbacts_cancel:|fbacts_exit)"))

    # Admin commands
    app.add_handler(CommandHandler("live", cmd_live))
    app.add_handler(CommandHandler("rcv", cmd_rcv))
    app.add_handler(CommandHandler("rcvall", cmd_rcvall))
    app.add_handler(CommandHandler("resetsub", cmd_resetsub))
    app.add_handler(CommandHandler("add", cmd_add))
    app.add_handler(CommandHandler("rm", cmd_rm))
    app.add_handler(CommandHandler("msg", cmd_msg))
    app.add_handler(CommandHandler("cast", cmd_cast))
    app.add_handler(CommandHandler("userinfo", cmd_userinfo))
    app.add_handler(CommandHandler("stp", cmd_stp))
    app.add_handler(CommandHandler("ldset", cmd_ldset))
    app.add_handler(CommandHandler("ldauto", cmd_ldauto))
    app.add_handler(CommandHandler("ldoff", cmd_ldoff))
    app.add_handler(CommandHandler("acts", cmd_acts))
    app.add_handler(CommandHandler("list", cmd_list))
    app.add_handler(CommandHandler("checktasks", cmd_checktasks))
    app.add_handler(CommandHandler("refreshtasks", cmd_refreshtasks))
    app.add_handler(CommandHandler("botoff", cmd_botoff))
    app.add_handler(CommandHandler("boton", cmd_boton))
    app.add_handler(CommandHandler("on2fa6h", cmd_on2fa6h))
    app.add_handler(CommandHandler("off2fa6h", cmd_off2fa6h))
    app.add_handler(CommandHandler("on2fa1h", cmd_on2fa1h))
    app.add_handler(CommandHandler("off2fa1h", cmd_off2fa1h))
    app.add_handler(CommandHandler("stats", cmd_stats))
    app.add_handler(CommandHandler("rmreview", cmd_rmreview))
    app.add_handler(CommandHandler("apr", cmd_apr))
    app.add_handler(CommandHandler("cmd", cmd_cmd))
    # FB commands
    app.add_handler(CommandHandler("fbon", cmd_fbon))
    app.add_handler(CommandHandler("fboff", cmd_fboff))
    app.add_handler(CommandHandler("fbstp", cmd_fbstp))
    app.add_handler(CommandHandler("fblive", cmd_fblive))
    app.add_handler(CommandHandler("fbrcv", cmd_fbrcv))
    app.add_handler(CommandHandler("fbrcvall", cmd_fbrcvall))
    app.add_handler(CommandHandler("fbacts", cmd_fbacts))

    app.add_error_handler(error_handler)

    logger.info("Bot is starting (polling)…")
    app.run_polling(allowed_updates=Update.ALL_TYPES, drop_pending_updates=True)


if __name__ == "__main__":
    main()
